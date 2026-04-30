from __future__ import annotations

import re
from pathlib import Path

import numpy as np
from scipy.interpolate import interp1d
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

#  STYLING

_FONT_HDR   = Font(name="Arial", size=10, bold=True, color="FFFFFF")
_FONT_DATA  = Font(name="Arial", size=10)
_FONT_TITLE = Font(name="Arial", size=11, bold=True, color="FFFFFF")
_ALIGN_C    = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_L    = Alignment(horizontal="left", vertical="center")
_BORDER     = Border(
    bottom=Side(style="thin", color="DFE6E9"),
    right=Side(style="thin", color="DFE6E9"),
)
_FILL_ALT   = PatternFill("solid", fgColor="F8FAF0")
_NUM_PT     = "0.0000"
_NUM_AN     = "0.000000"

_COL = {
    "Angles":  "1ABC9C",
    "Moments": "27AE60",
    "Powers":  "E74C3C",
    "Forces":  "5D6D7E",
    "GRF":     "884EA0",
    "EMG":     "2D3436",
    "summary": "9CBE20",
}

_N_NORM = 101  # 0 % … 100 % gait cycle

#  CELL HELPERS

def _section(ws, row, c0, c1, text, color="2D3436"):
    fill = PatternFill("solid", fgColor=color)
    for ci in range(c0, c1 + 1):
        ws.cell(row=row, column=ci).fill = fill
    c = ws.cell(row=row, column=c0, value=text)
    c.font = _FONT_TITLE
    c.alignment = _ALIGN_L
    ws.row_dimensions[row].height = 28

def _hdr(ws, row, col, text, color="2D3436"):
    c = ws.cell(row=row, column=col, value=text)
    c.font = _FONT_HDR
    c.fill = PatternFill("solid", fgColor=color)
    c.alignment = _ALIGN_C
    c.border = _BORDER

def _dat(ws, row, col, value, alt=False, fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = _FONT_DATA
    c.alignment = _ALIGN_L
    c.border = _BORDER
    if alt:
        c.fill = _FILL_ALT
    if fmt:
        c.number_format = fmt

#  EXCEL READER

def _read_sample_rates(wb) -> tuple[float, float, float]:
    """
    Read point (gait) and analog (EMG/GRF) sample rates + data start offset.
    Looks in Demographics first (C3D extractor output), then Info as fallback.
    Keys: "Sample Rate Gait" → point rate, "Sample Rate EMG" → analog rate.
    GRF uses the same analog rate as EMG.

    Returns (pt_rate, an_rate, data_start_s).
    data_start_s = time in seconds of the first data sample (= C3D
    POINT.FIRST_FRAME / point_rate). Recognises both "Data Start Offset"
    (current C3D extractor) and the legacy "Data Start Time" label.

    Falls back to deriving the offset from "First Frame" + "Point Sample
    Rate" if no explicit offset row is found.
    If nothing usable is found, defaults to 0.0.
    """
    pt, an, offset = 100.0, 1000.0, 0.0
    first_frame: float | None = None     # for fallback derivation

    for sheet_name in ("Demographics", "Info"):
        if sheet_name not in wb.sheetnames:
            continue
        for row in wb[sheet_name].iter_rows(min_row=2, max_col=2,
                                             values_only=True):
            if row[0] is None:
                continue
            key = str(row[0]).strip().lower()
            val = str(row[1]).strip() if row[1] is not None else ""
            # Accept negative numbers (e.g. scale factor) but for the keys
            # we care about all values are non-negative — keep it simple.
            nums = re.findall(r"[\d.]+", val)
            if not nums:
                continue
            if "sample rate gait" in key or "point sample rate" in key:
                pt = float(nums[0])
            elif "sample rate emg" in key or "analog sample rate" in key:
                an = float(nums[0])
            elif ("data start offset" in key
                  or "data start time" in key):
                # Explicit offset wins — use as-is.
                offset = float(nums[0])
            elif "first frame" in key:
                # Remember for fallback if no explicit offset row exists.
                try:
                    first_frame = float(nums[0])
                except ValueError:
                    pass

    # Fallback: derive offset from first_frame / pt_rate when no explicit
    # "Data Start Offset" row was present (older C3D-extractor outputs).
    # "First Frame" in the Info sheet is 1-indexed (matches C3D viewer
    # convention: frame 1 is at t=0). So offset = (first_frame - 1) / rate.
    if offset == 0.0 and first_frame is not None and pt > 0 and first_frame >= 1:
        offset = (first_frame - 1) / pt

    return pt, an, offset

def _read_events(wb) -> list[dict]:
    """Read events: [{context, label, time_s}, …] sorted by time."""
    events = []
    if "Events" not in wb.sheetnames:
        return events
    for row in wb["Events"].iter_rows(min_row=3, max_col=4, values_only=True):
        if row[0] is None or row[1] is None:
            continue
        try:
            events.append({
                "context": str(row[1]).strip(),
                "label":   str(row[2]).strip(),
                "time_s":  float(row[3]),
            })
        except (TypeError, ValueError, IndexError):
            continue
    events.sort(key=lambda e: e["time_s"])
    return events

def _read_channel_sheet(wb, name: str
                        ) -> tuple[list[str], list[list[float]], np.ndarray]:
    """
    Read a channel sheet.
    Returns (headers, data_columns, time_array).
    - headers: channel names from row 2 (columns B+)
    - data_columns: list of lists, one per channel
    - time_array: numpy array from column A (Frame or Time)
    Row 1 = section header, Row 2 = column headers, Row 3+ = data.
    Returns empty lists if the sheet doesn't exist or has no data.
    """
    if name not in wb.sheetnames:
        return [], [], np.array([])
    rows = list(wb[name].iter_rows(values_only=True))
    if len(rows) < 3:
        return [], [], np.array([])

    header_row = rows[1]
    headers = [str(h).strip() for h in header_row[1:] if h is not None]
    n_cols = len(headers)
    if n_cols == 0:
        return [], [], np.array([])

    time_vals: list[float] = []
    columns: list[list[float]] = [[] for _ in range(n_cols)]
    for row in rows[2:]:
        if row[0] is None:
            break
        try:
            time_vals.append(float(row[0]))
        except (TypeError, ValueError):
            time_vals.append(np.nan)
        for ci in range(n_cols):
            val = row[ci + 1] if ci + 1 < len(row) else None
            try:
                columns[ci].append(float(val))
            except (TypeError, ValueError):
                columns[ci].append(np.nan)

    return headers, columns, np.array(time_vals, dtype=float)

#  GAIT CYCLE DETECTION

def _find_gait_cycles(events: list[dict]) -> dict[str, list[dict]]:
    """
    A gait cycle = Foot Strike → next Foot Strike of the same side.
    Returns {"Right": [{start_s, end_s, fo_s}, …], "Left": …}
    """
    cycles: dict[str, list[dict]] = {"Right": [], "Left": []}

    for side in ("Right", "Left"):
        fs = [e["time_s"] for e in events
              if e["context"] == side and "strike" in e["label"].lower()]
        fo = [e["time_s"] for e in events
              if e["context"] == side and "off" in e["label"].lower()]

        for i in range(len(fs) - 1):
            t0, t1 = fs[i], fs[i + 1]
            t_fo = next((t for t in fo if t0 < t < t1), None)
            cycles[side].append({"start_s": t0, "end_s": t1, "fo_s": t_fo})

    return cycles

#  STRIDE EXTRACTION + NORMALISATION

def _extract_segment(col: list[float], t0: float, t1: float,
                     time_col: np.ndarray, rate: float) -> np.ndarray:
    """
    Extract samples between t0 and t1 using the actual time column.

    For point data:  time_col = frame_numbers → converted to seconds via /rate
    For analog data: time_col = already in seconds

    Uses np.searchsorted on the time column for precise index mapping,
    independent of sample rate rounding or first-frame offsets.
    """
    n = len(col)
    if n == 0:
        return np.array([])

    # Build a seconds-based time array from the column A data
    # Point sheets store frame numbers (0, 1, 2, …) → divide by rate
    # Analog sheets store time in seconds (0.0, 0.000667, …) → use as-is
    if len(time_col) > 1:
        # Detect: if values look like integer frame numbers (0,1,2,… or
        # starting at some offset like 200,201,202,…) — the step is ~1.0
        # For time-based columns the step is 1/rate (e.g. 0.001 for 1kHz)
        step = abs(time_col[1] - time_col[0])
        is_frames = (step > 0.5 and step < 1.5)  # step ≈ 1.0 → frames
        if is_frames:
            time_sec = time_col[:n] / rate
        else:
            time_sec = time_col[:n]
    else:
        time_sec = time_col[:n] / rate

    # Find indices via searchsorted (robust against rate mismatches)
    i0 = int(np.searchsorted(time_sec, t0, side="left"))
    i1 = int(np.searchsorted(time_sec, t1, side="right"))

    i0 = max(0, i0)
    i1 = min(n, i1)

    if i0 >= i1:
        return np.array([])
    return np.array(col[i0:i1], dtype=float)

def _normalise(data: np.ndarray, n: int = _N_NORM) -> np.ndarray:
    if len(data) < 2:
        return np.full(n, np.nan)
    mask = ~np.isnan(data)
    if mask.sum() < 2:
        return np.full(n, np.nan)
    x_orig = np.linspace(0, 100, len(data))
    x_norm = np.linspace(0, 100, n)
    kind = "cubic" if mask.sum() >= 4 else "linear"
    try:
        f = interp1d(x_orig[mask], data[mask], kind=kind,
                     bounds_error=False, fill_value="extrapolate")
        return f(x_norm)
    except Exception:
        return np.full(n, np.nan)

def _channel_side(header: str) -> str | None:
    """
    Detect which side a channel belongs to from its header name.
    Returns "R", "L", or None (no side / both sides).

    Uses known gait-model body-segment prefixes to avoid false positives
    like "LateralForce" → "L" or "RangeOfMotion" → "R".

    Patterns matched:
      "LHipAngles …"       → "L"    (L + known segment)
      "RKneeMoment …"      → "R"    (R + known segment)
      "EMG_LGastrocMed …"  → "L"    (EMG_ prefix then L)
      "EMG_RTibAnt …"      → "R"    (EMG_ prefix then R)
      "LateralForce …"     → None   (not a body segment)
      "Fx1 (N)"            → None   (force plate, no side)
    """
    h = header.strip()

    # Known body-segment prefixes used by Plug-in Gait / CGM / OLGA models
    _SEGMENTS = (
        "Hip", "Knee", "Ankle", "Pelvis", "Foot", "Toe",
        "Femur", "Tibia", "Shank", "Thigh",
        "Shoulder", "Elbow", "Wrist", "Hand",
        "Trunk", "Spine", "Thorax", "Head", "Neck",
        "GRF", "Ground",
        # Common combined labels
        "Normalised", "Abs",
    )

    # Check direct L/R prefix: "LHip…", "RKnee…"
    if len(h) >= 2 and h[0] in ("L", "R") and h[1].isupper():
        rest = h[1:]  # e.g. "HipAngles  [X / Sagittal] (deg)"
        # Must start with a known body segment
        if any(rest.startswith(seg) for seg in _SEGMENTS):
            return h[0]

    # Check EMG_L / EMG_R pattern
    for prefix in ("EMG_", "emg_", "EMG.", "emg."):
        if h.startswith(prefix) and len(h) > len(prefix):
            ch = h[len(prefix)]
            if ch in ("L", "R"):
                return ch

    return None

def _filter_side(headers: list[str], columns: list,
                 side: str) -> tuple[list[str], list]:
    """
    Keep only channels matching the given side.
    Right stride → R-channels + neutral channels.
    Left stride  → L-channels + neutral channels.
    """
    keep = "R" if side == "Right" else "L"
    filt_h, filt_c = [], []
    for h, c in zip(headers, columns):
        ch_side = _channel_side(h)
        if ch_side is None or ch_side == keep:
            filt_h.append(h)
            filt_c.append(c)
    return filt_h, filt_c

#  SHEET WRITERS

def _write_summary(wb, cycles, pt_rate, an_rate, data_offset, source):
    ws = wb.create_sheet("Summary")
    ws.sheet_properties.tabColor = _COL["summary"]

    _section(ws, 1, 1, 7,
             f"Stride Analysis  —  {source}", _COL["summary"])

    _dat(ws, 2, 1, "Point Rate (Hz)");       _dat(ws, 2, 2, pt_rate)
    _dat(ws, 3, 1, "Analog Rate (Hz)");      _dat(ws, 3, 2, an_rate)
    _dat(ws, 4, 1, "Normalisation Points");  _dat(ws, 4, 2, _N_NORM)
    _dat(ws, 5, 1, "Data Start Offset (s)"); _dat(ws, 5, 2, round(data_offset, 4))

    row = 7
    for side in ("Right", "Left"):
        sc = cycles[side]
        _section(ws, row, 1, 7,
                 f"{side}  —  {len(sc)} stride(s)", _COL["summary"])
        row += 1
        for ci, h in enumerate(("Stride", "Start (s)", "End (s)",
                                "Duration (s)", "Foot Off (s)",
                                "Stance %", "Swing %"), 1):
            _hdr(ws, row, ci, h, _COL["summary"])
        row += 1

        for si, cyc in enumerate(sc, 1):
            dur = cyc["end_s"] - cyc["start_s"]
            alt = si % 2 == 0
            _dat(ws, row, 1, f"S{si}", alt)
            _dat(ws, row, 2, round(cyc["start_s"], 4), alt, "0.0000")
            _dat(ws, row, 3, round(cyc["end_s"], 4), alt, "0.0000")
            _dat(ws, row, 4, round(dur, 4), alt, "0.0000")
            if cyc["fo_s"] is not None:
                stance = (cyc["fo_s"] - cyc["start_s"]) / dur * 100
                _dat(ws, row, 5, round(cyc["fo_s"], 4), alt, "0.0000")
                _dat(ws, row, 6, round(stance, 1), alt, "0.0")
                _dat(ws, row, 7, round(100 - stance, 1), alt, "0.0")
            else:
                for c in (5, 6, 7):
                    _dat(ws, row, c, "—", alt)
            row += 1
        row += 1

    for c in "ABCDEFG":
        ws.column_dimensions[c].width = 15

def _write_raw_point(wb, name, color, headers, stride_cols, rate):
    ws = wb.create_sheet(name)
    ws.sheet_properties.tabColor = color
    n = max((len(c) for c in stride_cols), default=0)

    _section(ws, 1, 1, len(headers) + 1,
             f"{name}  —  {n} frames @ {rate:.0f} Hz", color)

    _hdr(ws, 2, 1, "Frame", color)
    ws.column_dimensions["A"].width = 8
    for ci, h in enumerate(headers, 2):
        _hdr(ws, 2, ci, h, color)
        ws.column_dimensions[get_column_letter(ci)].width = 16
    ws.row_dimensions[2].height = 40
    ws.freeze_panes = "B3"

    for fi in range(n):
        row = fi + 3
        alt = fi % 2 == 0
        _dat(ws, row, 1, fi, alt, "0")
        for ci, col in enumerate(stride_cols, 2):
            v = col[fi] if fi < len(col) else None
            if v is not None and not np.isnan(v):
                _dat(ws, row, ci, round(float(v), 5), alt, _NUM_PT)
            else:
                _dat(ws, row, ci, None, alt)

def _write_norm_point(wb, name, color, headers, stride_cols_norm):
    ws = wb.create_sheet(name)
    ws.sheet_properties.tabColor = color

    _section(ws, 1, 1, len(headers) + 1,
             f"{name}  —  normalised to {_N_NORM} points (0–100 % GC)", color)

    _hdr(ws, 2, 1, "% GC", color)
    ws.column_dimensions["A"].width = 8
    for ci, h in enumerate(headers, 2):
        _hdr(ws, 2, ci, h, color)
        ws.column_dimensions[get_column_letter(ci)].width = 16
    ws.row_dimensions[2].height = 40
    ws.freeze_panes = "B3"

    for pct in range(_N_NORM):
        row = pct + 3
        alt = pct % 2 == 0
        _dat(ws, row, 1, pct, alt, "0")
        for ci, col in enumerate(stride_cols_norm, 2):
            v = col[pct]
            if not np.isnan(v):
                _dat(ws, row, ci, round(float(v), 5), alt, _NUM_PT)
            else:
                _dat(ws, row, ci, None, alt)

def _write_norm_mean(wb, name, color, headers, all_strides_norm, n_strides):
    ws = wb.create_sheet(name)
    ws.sheet_properties.tabColor = color

    _section(ws, 1, 1, len(headers) + 1,
             f"{name}  —  mean of {n_strides} stride(s)", color)

    _hdr(ws, 2, 1, "% GC", color)
    ws.column_dimensions["A"].width = 8
    for ci, h in enumerate(headers, 2):
        _hdr(ws, 2, ci, h, color)
        ws.column_dimensions[get_column_letter(ci)].width = 16
    ws.row_dimensions[2].height = 40
    ws.freeze_panes = "B3"

    for pct in range(_N_NORM):
        row = pct + 3
        alt = pct % 2 == 0
        _dat(ws, row, 1, pct, alt, "0")
        for ci, ch_strides in enumerate(all_strides_norm, 2):
            vals = [ch_strides[si][pct] for si in range(n_strides)]
            vals = [v for v in vals if not np.isnan(v)]
            if vals:
                _dat(ws, row, ci, round(float(np.mean(vals)), 5), alt, _NUM_PT)
            else:
                _dat(ws, row, ci, None, alt)

def _write_raw_analog(wb, name, color, headers, stride_cols, rate):
    ws = wb.create_sheet(name)
    ws.sheet_properties.tabColor = color
    n = max((len(c) for c in stride_cols), default=0)

    _section(ws, 1, 1, len(headers) + 1,
             f"{name}  —  {n:,} samples @ {rate:.0f} Hz", color)

    _hdr(ws, 2, 1, "Time (s)", color)
    ws.column_dimensions["A"].width = 12
    for ci, h in enumerate(headers, 2):
        _hdr(ws, 2, ci, h, color)
        ws.column_dimensions[get_column_letter(ci)].width = 16
    ws.row_dimensions[2].height = 40
    ws.freeze_panes = "B3"

    for si in range(n):
        row = si + 3
        alt = si % 2 == 0
        _dat(ws, row, 1, round(si / rate, 6), alt, _NUM_AN)
        for ci, col in enumerate(stride_cols, 2):
            v = col[si] if si < len(col) else None
            if v is not None and not np.isnan(v):
                _dat(ws, row, ci, round(float(v), 6), alt, _NUM_AN)
            else:
                _dat(ws, row, ci, None, alt)

#  MAIN PUBLIC FUNCTION

def analyse_strides(
    input_path:  str | Path,
    output_path: str | Path,
    status_cb=None,
) -> dict:
    """
    Read a C3D-generated Excel, extract strides, write stride-analysis workbook.

    Output sheet order (grouped by stride):
      Summary
      S1_R_Angles_Raw, S1_R_Angles_Norm, S1_R_Moments_Raw, …, S1_R_EMG, S1_R_GRF
      S1_L_Angles_Raw, S1_L_Angles_Norm, …
      S2_R_…, S2_L_…, …
      Mean_R_Angles, Mean_R_Moments, …  (mean across all strides)
      Mean_L_Angles, Mean_L_Moments, …

    Missing channel sheets are silently skipped.
    EMG/GRF use the analog sample rate for stride boundaries.
    """
    input_path  = Path(input_path)
    output_path = Path(output_path)

    def upd(msg):
        if status_cb:
            status_cb(msg)

    # 1. Load
    upd(f"Loading {input_path.name} …")
    wb_in = openpyxl.load_workbook(str(input_path), data_only=True,
                                    read_only=True)

    pt_rate, an_rate, data_offset = _read_sample_rates(wb_in)
    upd(f"  Point rate: {pt_rate} Hz  |  Analog rate: {an_rate} Hz")
    if data_offset > 0:
        upd(f"  Data start offset: {data_offset:.4f} s")

    events = _read_events(wb_in)
    upd(f"  Events: {len(events)}")
    if not events:
        wb_in.close()
        raise ValueError("No events found — cannot detect gait cycles.")

    # 2. Detect gait cycles
    cycles = _find_gait_cycles(events)
    n_r, n_l = len(cycles["Right"]), len(cycles["Left"])
    upd(f"  Strides: {n_r} Right, {n_l} Left")
    if n_r == 0 and n_l == 0:
        wb_in.close()
        raise ValueError("No complete gait cycles. Need ≥ 2 Foot Strike "
                         "events per side.")

    # 3. Pre-read all available channel sheets
    point_names  = ["Angles", "Moments", "Powers", "Forces"]
    analog_names = ["EMG", "GRF"]

    # channel_data[name] = (headers, columns, time_col)
    channel_data: dict[str, tuple[list[str], list[list[float]], np.ndarray]] = {}

    for name in point_names + analog_names:
        hdrs, cols, time_col = _read_channel_sheet(wb_in, name)
        if hdrs and cols:
            channel_data[name] = (hdrs, cols, time_col)
            upd(f"  {name}: {len(hdrs)} channels, {len(cols[0]):,} samples")
        else:
            upd(f"  {name}: not found or empty — skipping")

    wb_in.close()

    if not channel_data:
        raise ValueError("No channel data found in any sheet.")

    # 4. Build output workbook
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)
    sheets = 0

    _write_summary(wb_out, cycles, pt_rate, an_rate, data_offset,
                   input_path.name)
    sheets += 1

    # Convert event times from absolute to data-relative
    # (subtract C3D first-frame offset so indices match Excel row 0 = time 0)
    if data_offset > 0:
        upd(f"  Adjusting event times by -{data_offset:.4f} s offset …")
    for side in cycles:
        for cyc in cycles[side]:
            cyc["start_s"] -= data_offset
            cyc["end_s"]   -= data_offset
            if cyc["fo_s"] is not None:
                cyc["fo_s"] -= data_offset

    # Collect normalised data for mean calculation later
    # norm_bank[sheet_name][side] = {"hdrs": [...], "data": [[stride1, stride2, ...], ...]}
    norm_bank: dict[str, dict[str, dict]] = {}

    max_strides = max(n_r, n_l)

    # 5. Per-stride sheets (S1, S2, …)
    for si in range(max_strides):
        for side in ("Right", "Left"):
            tag = "R" if side == "Right" else "L"
            side_cycles = cycles[side]
            if si >= len(side_cycles):
                continue
            cyc = side_cycles[si]

            upd(f"Stride {si+1} {side} …")

            # — Point-based channels: raw + norm (POINT rate) ─────────
            for name in point_names:
                if name not in channel_data:
                    continue
                hdrs_all, cols_all, time_col = channel_data[name]
                color = _COL.get(name, "2D3436")

                # Filter: Right stride → R-channels only, Left → L-channels
                hdrs, cols = _filter_side(hdrs_all, cols_all, side)
                if not hdrs:
                    continue

                raw_cols  = [_extract_segment(c, cyc["start_s"],
                                              cyc["end_s"],
                                              time_col, pt_rate)
                             for c in cols]
                norm_cols = [_normalise(rc) for rc in raw_cols]

                # Raw sheet
                sname = f"S{si+1}_{tag}_{name}_Raw"
                _write_raw_point(wb_out, sname, color, hdrs,
                                 raw_cols, pt_rate)
                sheets += 1

                # Norm sheet
                sname = f"S{si+1}_{tag}_{name}_Norm"
                _write_norm_point(wb_out, sname, color, hdrs, norm_cols)
                sheets += 1

                # Bank for mean
                if name not in norm_bank:
                    norm_bank[name] = {}
                if side not in norm_bank[name]:
                    norm_bank[name][side] = {
                        "hdrs": hdrs,
                        "data": [[] for _ in hdrs],
                    }
                for ch_i, nc in enumerate(norm_cols):
                    norm_bank[name][side]["data"][ch_i].append(nc)

            # — Analog channels: raw only (ANALOG rate) ───────────────
            for name in analog_names:
                if name not in channel_data:
                    continue
                hdrs_all, cols_all, time_col = channel_data[name]
                color = _COL.get(name, "2D3436")

                # Filter: Right stride → R-channels only, Left → L-channels
                hdrs, cols = _filter_side(hdrs_all, cols_all, side)
                if not hdrs:
                    continue

                # KEY: use an_rate + time_col for correct stride boundaries
                raw_cols = [_extract_segment(c, cyc["start_s"],
                                             cyc["end_s"],
                                             time_col, an_rate)
                            for c in cols]

                sname = f"S{si+1}_{tag}_{name}"
                _write_raw_analog(wb_out, sname, color, hdrs,
                                  raw_cols, an_rate)
                sheets += 1

    # 6. Mean sheets (average of all normalised strides)
    for name in point_names:
        if name not in norm_bank:
            continue
        color = _COL.get(name, "2D3436")

        for side in ("Right", "Left"):
            tag = "R" if side == "Right" else "L"
            n_s = len(cycles[side])
            if n_s == 0 or side not in norm_bank[name]:
                continue

            bank = norm_bank[name][side]
            hdrs = bank["hdrs"]
            data = bank["data"]

            sname = f"Mean_{tag}_{name}"
            upd(f"  Writing {sname} ({n_s} strides) …")
            _write_norm_mean(wb_out, sname, color, hdrs, data, n_s)
            sheets += 1

    # 7. Save
    upd(f"Saving {output_path.name} …")
    wb_out.save(str(output_path))
    upd(f"✓ Done — {sheets} sheets written.")

    return {
        "right_strides":  n_r,
        "left_strides":   n_l,
        "sheets_written": sheets,
    }

#  PySide6 WORKER + TAB

try:
    from PySide6.QtCore import Qt, Signal, QThread, QTimer
    from PySide6.QtWidgets import (
        QWidget, QVBoxLayout, QHBoxLayout, QGroupBox, QLabel,
        QPushButton, QProgressBar, QFileDialog, QMessageBox,
        QTableWidget, QTableWidgetItem, QAbstractItemView,
        QScrollArea,
    )
    from PySide6.QtGui import QColor
    _HAS_PYSIDE = True
except ImportError:
    _HAS_PYSIDE = False

if _HAS_PYSIDE:

    from GaitSharing_config import PALETTE

    class StrideWorker(QThread):
        progress    = Signal(str, int)           # message, percent
        file_status = Signal(str, str, str)      # path, text, type
        log_msg     = Signal(str)
        finished    = Signal(int, list)          # success_count, errors

        def __init__(self, files: list, out_dir):
            super().__init__()
            self.files   = files
            self.out_dir = out_dir

        def run(self):
            total   = len(self.files)
            success = 0
            errors  = []

            for i, xlsx_path in enumerate(self.files):
                self.progress.emit(
                    f"[{i+1}/{total}]  {xlsx_path.name}",
                    int(100 * i / total))
                self.file_status.emit(str(xlsx_path), "Running…", "running")

                out_path = self.out_dir / f"{xlsx_path.stem}_strides.xlsx"

                def _cb(msg, p=xlsx_path):
                    self.log_msg.emit(f"    {msg}")

                try:
                    self.log_msg.emit(
                        f"\n{'─'*50}\n{xlsx_path.name}\n{'─'*50}")
                    result = analyse_strides(xlsx_path, out_path, status_cb=_cb)
                    success += 1
                    n_r = result["right_strides"]
                    n_l = result["left_strides"]
                    n_s = result["sheets_written"]
                    self.file_status.emit(
                        str(xlsx_path),
                        f"✓ R:{n_r} L:{n_l} ({n_s} sheets)", "ok")
                    self.log_msg.emit(
                        f"  ✓  {xlsx_path.name}  →  {out_path.name}")
                except Exception as exc:
                    import traceback
                    tb = traceback.format_exc()
                    errors.append((xlsx_path.name, str(exc)))
                    self.file_status.emit(
                        str(xlsx_path), f"✗ {exc}", "err")
                    self.log_msg.emit(
                        f"  ✗  {xlsx_path.name}:  {exc}\n{tb}")

            self.finished.emit(success, errors)

    class StrideAnalysisTab(QWidget):

        def __init__(self, parent=None):
            super().__init__(parent)
            self._files: list[Path] = []
            self._out_dir: Path | None = None
            self._worker = None
            self._build()

        def _build(self):
            layout = QVBoxLayout(self)
            layout.setContentsMargins(0, 0, 0, 0)
            layout.setSpacing(0)

            from GaitSharing_ui import (make_page_header, make_accent_btn,
                                        make_console_log)

            layout.addWidget(make_page_header(
                "Stride Analysis",
                "Extract individual gait cycles, normalise to 0–100 % GC"))

            content = QWidget()
            clay = QVBoxLayout(content)
            clay.setContentsMargins(16, 12, 16, 12)
            clay.setSpacing(10)

            # File list
            file_group = QGroupBox("Input Excel Files (from C3D Extractor)")
            fglay = QVBoxLayout(file_group)

            btn_row = QHBoxLayout()
            add_btn = make_accent_btn("+ Add Excel Files")
            add_btn.clicked.connect(self._add_files)
            btn_row.addWidget(add_btn)
            rm_btn = QPushButton("Remove Selected")
            rm_btn.clicked.connect(self._remove_sel)
            btn_row.addWidget(rm_btn)
            clr_btn = QPushButton("Clear All")
            clr_btn.clicked.connect(self._clear)
            btn_row.addWidget(clr_btn)
            btn_row.addStretch()
            fglay.addLayout(btn_row)

            self.file_table = QTableWidget()
            self.file_table.setColumnCount(3)
            self.file_table.setHorizontalHeaderLabels(
                ["Filename", "Folder", "Status"])
            self.file_table.setAlternatingRowColors(True)
            self.file_table.setSelectionBehavior(
                QAbstractItemView.SelectRows)
            self.file_table.setSelectionMode(
                QAbstractItemView.ExtendedSelection)
            self.file_table.verticalHeader().setVisible(False)
            self.file_table.setEditTriggers(
                QAbstractItemView.NoEditTriggers)
            self.file_table.horizontalHeader().setStretchLastSection(True)
            self.file_table.setColumnWidth(0, 220)
            self.file_table.setColumnWidth(1, 360)
            self.file_table.setMaximumHeight(140)
            fglay.addWidget(self.file_table)

            self._file_count_lbl = QLabel("No files selected.")
            self._file_count_lbl.setStyleSheet(
                f"color: {PALETTE['text_muted']}; font-size: 12px;")
            fglay.addWidget(self._file_count_lbl)
            clay.addWidget(file_group)

            # Output folder
            out_group = QGroupBox(
                "Output Folder (one _strides.xlsx per input file)")
            oglay = QHBoxLayout(out_group)
            self._out_lbl = QLabel("No folder selected.")
            self._out_lbl.setStyleSheet(
                f"color: {PALETTE['text_muted']};")
            oglay.addWidget(self._out_lbl, stretch=1)
            brw_btn = QPushButton("Browse…")
            brw_btn.clicked.connect(self._browse_out)
            oglay.addWidget(brw_btn)
            clay.addWidget(out_group)

            # Info label
            info = QLabel(
                "ℹ  Reads Events + sample rates from the Excel, "
                "splits all available channels into strides.\n"
                "     Missing sheets (Angles, EMG, etc.) are "
                "silently skipped.  EMG/GRF use the analog sample "
                "rate.")
            info.setWordWrap(True)
            info.setStyleSheet(
                f"color: {PALETTE['text_muted']}; font-size: 12px; "
                f"padding: 4px 0;")
            clay.addWidget(info)

            # Run button
            self._run_btn = make_accent_btn("▶   Analyse Strides")
            self._run_btn.clicked.connect(self._run)
            clay.addWidget(self._run_btn, alignment=Qt.AlignLeft)

            # Missing deps warning
            missing = []
            try:
                import scipy      # noqa: F401
            except ImportError:
                missing.append("scipy")
            if missing:
                warn = QLabel(
                    "⚠  Missing: " + ", ".join(missing) +
                    "   →   pip install " + " ".join(missing))
                warn.setStyleSheet(
                    f"color: {PALETTE['warning']}; font-weight: 700;")
                clay.addWidget(warn)

            # Progress
            prog_group = QGroupBox("Progress")
            pglay = QVBoxLayout(prog_group)
            self._status_lbl = QLabel("Ready.")
            pglay.addWidget(self._status_lbl)
            self._pbar = QProgressBar()
            self._pbar.setMaximum(100)
            pglay.addWidget(self._pbar)
            clay.addWidget(prog_group)

            # Log
            log_group = QGroupBox("Log")
            lglay = QVBoxLayout(log_group)
            self._log_widget = make_console_log()
            self._log_widget.setMinimumHeight(100)
            lglay.addWidget(self._log_widget)
            clay.addWidget(log_group, stretch=1)

            # Wrap in scroll area for small screens (14")
            scroll = QScrollArea()
            scroll.setWidgetResizable(True)
            scroll.setWidget(content)
            scroll.setStyleSheet("QScrollArea { border: none; }")
            layout.addWidget(scroll)

        # Helpers

        def _log(self, msg: str):
            self._log_widget.append(msg)

        def _refresh_file_list(self):
            self.file_table.setRowCount(len(self._files))
            for i, p in enumerate(self._files):
                self.file_table.setItem(
                    i, 0, QTableWidgetItem(p.name))
                self.file_table.setItem(
                    i, 1, QTableWidgetItem(str(p.parent)))
                item = QTableWidgetItem("Pending")
                item.setForeground(QColor(PALETTE["text_muted"]))
                self.file_table.setItem(i, 2, item)

            n = len(self._files)
            self._file_count_lbl.setText(
                f"{n} file{'s' if n != 1 else ''} selected."
                if n else "No files selected.")

        def _set_file_status(self, path_str, text, status_type):
            for row in range(self.file_table.rowCount()):
                item0 = self.file_table.item(row, 0)
                item1 = self.file_table.item(row, 1)
                if item0 and item1:
                    full = str(Path(item1.text()) / item0.text())
                    if full == path_str:
                        status_item = QTableWidgetItem(text)
                        color_map = {
                            "running": PALETTE["warning"],
                            "ok":      PALETTE["success"],
                            "err":     "#C42B1C",
                        }
                        status_item.setForeground(
                            QColor(color_map.get(
                                status_type, PALETTE["text"])))
                        self.file_table.setItem(row, 2, status_item)
                        break

        def _add_files(self):
            files, _ = QFileDialog.getOpenFileNames(
                self, "Select Excel Files (from C3D Extractor)",
                filter="Excel files (*.xlsx);;All files (*.*)")
            for raw in files:
                p = Path(raw)
                if p not in self._files:
                    self._files.append(p)
            self._refresh_file_list()

        def _remove_sel(self):
            rows = sorted(
                set(idx.row()
                    for idx in self.file_table.selectedIndexes()),
                reverse=True)
            for row in rows:
                if row < len(self._files):
                    self._files.pop(row)
            self._refresh_file_list()

        def _clear(self):
            self._files.clear()
            self._refresh_file_list()

        def _browse_out(self):
            d = QFileDialog.getExistingDirectory(
                self, "Select Output Folder")
            if d:
                self._out_dir = Path(d)
                self._out_lbl.setText(str(self._out_dir))

        def _run(self):
            if not self._files:
                QMessageBox.warning(
                    self, "No Files",
                    "Add at least one .xlsx file.")
                return
            if not self._out_dir:
                QMessageBox.critical(
                    self, "No Output Folder",
                    "Select an output folder.")
                return

            self._run_btn.setEnabled(False)
            self._pbar.setValue(0)

            self._worker = StrideWorker(
                self._files[:], self._out_dir)
            self._worker.progress.connect(
                lambda msg, pct: (
                    self._status_lbl.setText(msg),
                    self._pbar.setValue(pct)))
            self._worker.file_status.connect(self._set_file_status)
            self._worker.log_msg.connect(self._log)
            self._worker.finished.connect(self._on_finished)
            self._worker.start()

        def _on_finished(self, success, errors):
            total = len(self._files)
            self._pbar.setValue(100)
            self._status_lbl.setText(
                f"Done — {success}/{total} files, "
                f"{len(errors)} error(s).")
            self._run_btn.setEnabled(True)

            if errors:
                QMessageBox.warning(
                    self, "Done",
                    f"{success} OK, {len(errors)} failed.\n\n" +
                    "\n".join(f"• {n}: {e}"
                              for n, e in errors[:5]))
            else:
                QMessageBox.information(
                    self, "Done",
                    f"✓  {success} file(s) saved to:\n"
                    f"{self._out_dir}")

#  CLI

if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("Usage: python GaitSharing_strides.py <input.xlsx> [output.xlsx]")
        sys.exit(1)
    inp = Path(sys.argv[1])
    out = Path(sys.argv[2]) if len(sys.argv) > 2 else inp.with_name(
        f"{inp.stem}_strides{inp.suffix}")
    result = analyse_strides(inp, out, status_cb=print)
    print(f"\nResult: {result}")
