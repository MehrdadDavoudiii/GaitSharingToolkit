from __future__ import annotations
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None

# Column map (db_field → display_header)
EXPORT_COLS: list[tuple[str, str]] = [
    ("folder_name",     "Folder"),
    ("ganglabor_id",    "Gait Lab ID"),
    ("last_name",       "Last Name"),
    ("first_name",      "First Name"),
    ("birth_date",      "Birth Date"),
    ("gender",          "Gender"),
    ("exam_date",       "Exam Date"),
    ("diagnosis",       "Diagnosis"),
    ("gmfcs",           "GMFCS"),
    ("fms_5",           "FMS 5m"),
    ("fms_50",          "FMS 50m"),
    ("fms_500",         "FMS 500m"),
    ("fragestellung",   "Clinical Question / Fragestellung"),
    ("condition_left",  "Condition Left"),
    ("condition_right", "Condition Right"),
    ("measurements",    "Measurements"),
    ("emg_channels",    "EMG Channels"),
    ("model",           "Model"),
    ("examiner",        "Examiner"),
    ("import_date",     "Import Date"),
]

# Column widths (1-indexed, matching EXPORT_COLS order)
_WIDTHS = {
    1:14, 2:14, 3:16, 4:14, 5:12, 6:8, 7:12, 8:45, 9:8,
    10:8, 11:8, 12:8, 13:45, 14:18, 15:18, 16:22, 17:10,
    18:16, 19:18, 20:14,
}

def create_export_excel(subjects: list[dict], dest_path: Path) -> None:
    if openpyxl is None:
        raise RuntimeError("openpyxl is not installed — pip install openpyxl")

    HDR_FILL = PatternFill("solid", start_color="1A3A5C")
    HDR_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    ALT_FILL = PatternFill("solid", start_color="EEF3FB")
    NRM_FONT = Font(name="Calibri", size=10)
    BS       = Side(style="thin", color="CCCCCC")
    BDR      = Border(left=BS, right=BS, top=BS, bottom=BS)
    CTR_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
    TOP_WRAP = Alignment(vertical="top", wrap_text=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Subjects"

    # header row
    for ci, (_, label) in enumerate(EXPORT_COLS, 1):
        c = ws.cell(row=1, column=ci, value=label)
        c.font = HDR_FONT; c.fill = HDR_FILL
        c.alignment = CTR_WRAP; c.border = BDR
    ws.row_dimensions[1].height = 30

    # data rows
    for ri, subj in enumerate(subjects, 2):
        fill = ALT_FILL if ri % 2 == 0 else None
        for ci, (field, _) in enumerate(EXPORT_COLS, 1):
            c = ws.cell(row=ri, column=ci, value=subj.get(field) or "")
            c.font = NRM_FONT; c.border = BDR; c.alignment = TOP_WRAP
            if fill:
                c.fill = fill

    # column widths & freeze
    for ci, w in _WIDTHS.items():
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"

    # summary sheet
    ws2 = wb.create_sheet("Summary")
    for r, (label, val) in enumerate([
        ("Export Date",   datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Total Subjects", len(subjects)),
        ("Tool",           "Gait Sharing"),
    ], 1):
        ws2.cell(row=r, column=1, value=label).font = Font(bold=True, name="Calibri")
        ws2.cell(row=r, column=2, value=val)
    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 30

    wb.save(str(dest_path))
