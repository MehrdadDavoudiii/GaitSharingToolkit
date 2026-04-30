
from __future__ import annotations
import re, threading
from pathlib import Path

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QGroupBox,
    QLabel, QPushButton, QProgressBar, QTextEdit,
    QTableWidget, QTableWidgetItem, QHeaderView,
    QAbstractItemView, QFileDialog, QMessageBox,
    QCheckBox, QFrame, QScrollArea, QSplitter,
)
from PySide6.QtCore import Qt, QThread, Signal, QTimer
from PySide6.QtGui import QColor

try:
    import numpy as np
except ImportError:
    np = None

try:
    import ezc3d
except ImportError:
    ezc3d = None

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None

from GaitSharing_config import PALETTE

# optional PDF fallback (same parser used for import)
try:
    from GaitSharing_parser import (
        find_report_pdf, extract_text_from_pdf, parse_pdf_fields,
    )
    _PDF_PARSER_OK = True
except Exception:
    _PDF_PARSER_OK = False

#  KEYWORD CLASSIFIERS  (case-insensitive label matching)

def _lc(s: str) -> str:
    return s.lower()

# Point channel classifiers
def _is_angle(lbl: str)   -> bool: return "angle"   in _lc(lbl)
def _is_moment(lbl: str)  -> bool: return "moment"  in _lc(lbl)
def _is_power(lbl: str)   -> bool: return "power"   in _lc(lbl)
def _is_force_pt(lbl: str)-> bool: return "force"   in _lc(lbl)   # point-based forces

# Analog channel classifiers
def _is_emg(lbl: str)     -> bool: return "emg"     in _lc(lbl)
def _is_grf(lbl: str) -> bool:
    l = _lc(lbl)
    return ("grf" in l or "ground" in l or "force" in l or
            re.search(r"\bf[xyz]\d*\b|\bm[xyz]\d*\b|\bcop[xyz]?\d*\b", l) is not None)
def _is_fp_force(lbl: str) -> bool:  # kept for invert logic
    return _is_grf(lbl)

# Clinical PROCESSING keyword filter
_CLINICAL_KW = ["strength", "spastici", "function", "length", "width",
                "kraft", "spastik", "funktion", "lange", "länge", "breite",
                "spasm", "rom", "range"]

#  LOW-LEVEL HELPERS

def _safe_float(v) -> float | None:
    try:
        f = float(v)
        return None if (np is not None and np.isnan(f)) else f
    except Exception:
        return None

def _point_rate(c3d: dict) -> float:
    for path in (("header", "points", "frame_rate"),
                 ("parameters", "POINT", "RATE", "value")):
        try:
            node = c3d
            for k in path: node = node[k]
            v = float(node[0]) if isinstance(node, (list, tuple)) else float(node)
            if v > 0: return v
        except Exception: pass
    return 100.0

def _analog_rate(c3d: dict) -> float:
    for path in (("header", "analogs", "frame_rate"),
                 ("parameters", "ANALOG", "RATE", "value")):
        try:
            node = c3d
            for k in path: node = node[k]
            v = float(node[0]) if isinstance(node, (list, tuple)) else float(node)
            if v > 0: return v
        except Exception: pass
    return 1000.0

def _get_list(c3d: dict, *path) -> list:
    try:
        node = c3d
        for k in path: node = node[k]
        if isinstance(node, (list, tuple)): return list(node)
        if np is not None and isinstance(node, np.ndarray):
            return node.flatten().tolist()
        return [node]
    except Exception:
        return []

def _get_scalar(c3d: dict, *path):
    v = _get_list(c3d, *path)
    if not v: return None
    val = v[0]
    if isinstance(val, (list, tuple)): return val[0] if val else None
    return val

def _point_unit(c3d: dict) -> str:
    u = _get_scalar(c3d, "parameters", "POINT", "UNITS", "value")
    return str(u).strip() if u else ""

def _first_frame(c3d: dict) -> int:
    """
    Read the C3D header's FIRST_FRAME value.

    In Vicon/Nexus C3D files, recording often starts at a frame > 1 (e.g. 17),
    meaning the data array's first sample corresponds to that absolute frame
    number. Event times in the C3D EVENT block are absolute (referenced to
    frame 1 / time 0), so to align events with the data we need this offset.

    Returns 0 if the header doesn't expose it (i.e. data starts at frame 0/1).
    Tries (in order):
        c3d["header"]["points"]["first_frame"]
        c3d["parameters"]["POINT"]["FIRST_FRAME"]["value"]
        c3d["parameters"]["TRIAL"]["ACTUAL_START_FIELD"]["value"]
    """
    for path in (("header", "points", "first_frame"),
                 ("parameters", "POINT", "FIRST_FRAME", "value"),
                 ("parameters", "TRIAL", "ACTUAL_START_FIELD", "value")):
        try:
            node = c3d
            for k in path:
                node = node[k]
            if isinstance(node, (list, tuple)):
                node = node[0] if node else None
            if np is not None and isinstance(node, np.ndarray):
                node = node.flat[0] if node.size > 0 else None
            if node is None:
                continue
            v = int(node)
            if v >= 0:
                return v
        except Exception:
            pass
    return 0

def _last_frame(c3d: dict) -> int:
    for path in (("header", "points", "last_frame"),
                 ("parameters", "POINT", "LAST_FRAME", "value"),
                 ("parameters", "TRIAL", "ACTUAL_END_FIELD", "value")):
        try:
            node = c3d
            for k in path:
                node = node[k]
            if isinstance(node, (list, tuple)):
                node = node[0] if node else None
            if np is not None and isinstance(node, np.ndarray):
                node = node.flat[0] if node.size > 0 else None
            if node is None:
                continue
            v = int(node)
            if v > 0:
                return v
        except Exception:
            pass
    return 0

def _data_start_offset_s(c3d: dict) -> float:
    """
    Time (in seconds) of the first recorded frame, relative to trial start.

    Equals first_frame / point_rate.  Used by the stride extractor to shift
    absolute event times into data-relative time so indices line up with
    the Excel rows produced by the C3D extractor.
    """
    fr = _point_rate(c3d)
    ff = _first_frame(c3d)
    if fr <= 0:
        return 0.0
    return ff / fr

def _analog_units(c3d: dict, n_ch: int) -> list[str]:
    raw = _get_list(c3d, "parameters", "ANALOG", "UNITS", "value")
    units = [str(u).strip() for u in raw]
    # pad / trim to n_ch
    if len(units) < n_ch:
        units += [""] * (n_ch - len(units))
    return units[:n_ch]

#  DATA EXTRACTORS

def _extract_points_by_type(
    c3d: dict,
    classifier,          # callable(label) -> bool
) -> tuple[list[str], list[list]]:
    """
    Extract all POINT channels matching *classifier*.
    Returns (headers, columns).
    Headers: "{label} [{plane}] ({unit})"
    Columns: list of n_frames values per column (float or None)
    Planes: X=Sagittal, Y=Frontal, Z=Transverse
    """
    if np is None:
        return [], []
    try:
        pts    = c3d["data"]["points"]                        # (4, n_labels, n_frames)
        labels = [str(x).strip()
                  for x in c3d["parameters"]["POINT"]["LABELS"]["value"]]
        # Fixed units per channel type — ignore C3D POINT.UNITS (unreliable)
        def _unit_for(lbl):
            l = _lc(lbl)
            if "angle"  in l: return "deg"
            if "moment" in l: return "Nmm"
            if "power"  in l: return "W"
            if "force"  in l: return "N"
            return _point_unit(c3d)   # fallback to C3D stored unit

        planes = [("X / Sagittal", 0), ("Y / Frontal", 1), ("Z / Transverse", 2)]

        headers: list[str] = []
        columns: list[list] = []

        for li, lbl in enumerate(labels):
            if not classifier(lbl):
                continue
            unit_str = f" ({_unit_for(lbl)})"
            for plane_name, pi in planes:
                seg = pts[pi, li, :].astype(float)
                col = [None if np.isnan(v) else round(float(v), 5) for v in seg]
                headers.append(f"{lbl}  [{plane_name}]{unit_str}")
                columns.append(col)

        return headers, columns
    except Exception as exc:
        print(f"[_extract_points_by_type] {exc}")
        return [], []

def _extract_analogs_by_type(
    c3d:        dict,
    classifier,           # callable(label) -> bool
    invert:     bool = False,  # if True → channels that do NOT match any other class
) -> tuple[list[str], list[list]]:
    """
    Extract all ANALOG channels matching *classifier*.
    Returns (headers, columns).
    Headers: "{label} ({unit})"
    """
    if np is None:
        return [], []
    try:
        raw    = c3d["data"]["analogs"]          # (1, n_ch, n_samp) or (n_ch, n_samp)
        labels = [str(x).strip()
                  for x in c3d["parameters"]["ANALOG"]["LABELS"]["value"]]
        if raw.ndim == 3:
            raw = raw[0]
        n_ch, n_samp = raw.shape
        units = _analog_units(c3d, n_ch)

        headers: list[str] = []
        columns: list[list] = []

        for i, lbl in enumerate(labels[:n_ch]):
            if invert:
                # "Other" = doesn't match any named category
                match = (
                    _is_emg(lbl) or _is_grf(lbl) or _is_fp_force(lbl)
                )
                if match:
                    continue
            else:
                if not classifier(lbl):
                    continue

            unit   = units[i]
            u_str  = f" ({unit})" if unit else ""
            col    = [round(float(v), 6) for v in raw[i, :].tolist()]
            headers.append(f"{lbl}{u_str}")
            columns.append(col)

        return headers, columns
    except Exception as exc:
        print(f"[_extract_analogs_by_type] {exc}")
        return [], []

def read_events(c3d: dict) -> list[dict]:
    """All events sorted by time. Each: {context, label, time_s}."""
    events: list[dict] = []
    if np is None:
        return events
    try:
        ev       = c3d["parameters"]["EVENT"]
        contexts = [str(x).strip() for x in ev.get("CONTEXTS", {}).get("value", [])]
        labels   = [str(x).strip() for x in ev.get("LABELS",   {}).get("value", [])]
        raw      = np.array(ev.get("TIMES", {}).get("value", []), dtype=float)
        if raw.size == 0:
            return events
        # Row 1 (0-indexed) = seconds per MATLAB source: TIMES.DATA(2,:)
        if raw.ndim == 2 and raw.shape[0] >= 2:
            sec = raw[1, :]
        elif raw.ndim == 2:
            sec = raw[0, :]
        else:
            sec = raw
        n = min(len(contexts), len(labels), sec.size)
        for i in range(n):
            events.append({
                "context": contexts[i],
                "label":   labels[i],
                "time_s":  float(sec[i]),
            })
        events.sort(key=lambda e: e["time_s"])
    except Exception as exc:
        print(f"[read_events] {exc}")
    return events

def read_demographics_c3d(c3d: dict) -> dict[str, object]:

    result: dict[str, object] = {}

    def _read_param(group: str, key: str):
        """Helper to safely read a parameter from a specific group."""
        try:
            node = c3d["parameters"][group][key]
            v = None
            for leaf in ("value", "DATA"):
                if leaf in node:
                    v = node[leaf]
                    break
            if v is None: 
                return None
            
            if isinstance(v, (list, tuple)): 
                v = v[0] if v else None
            if np is not None and isinstance(v, np.ndarray):
                v = v.flat[0] if v.size > 0 else None
                
            if isinstance(v, bytes):
                v = v.decode('utf-8', errors='ignore')
                
            return v
        except Exception:
            return None

    gait_id = (_read_param("SUBJECTS", "NAMES") or 
               _read_param("SUBJECTS", "USED") or 
               _read_param("PROCESSING", "GanglaborID") or 
               _read_param("PROCESSING", "GaitLabID") or 
               _read_param("PROCESSING", "ID"))
    
    if gait_id is not None:
        result["Gait Lab ID"] = str(gait_id).strip()

    bm = _read_param("PROCESSING", "Bodymass")
    if bm is not None:
        try: 
            result["Bodymass (kg)"] = str(round(float(bm), 1))
        except Exception: 
            result["Bodymass (kg)"] = str(bm).strip()
    ht = _read_param("PROCESSING", "Height")
    if ht is not None:
         result["Height (mm)"] = str(ht).strip()

    age = _read_param("PROCESSING", "Age")
    if age is not None:
        try: 
            result["Age"] = str(round(float(age), 1))
        except Exception: 
            result["Age"] = str(age).strip()

    gen = _read_param("PROCESSING", "Gender")
    if gen is not None:
        result["Gender"] = str(gen).strip()

    Freq_gait = (_read_param("POINT", "RATE"))
    
    if Freq_gait is not None:
        result["Sample Rate Gait"] = str(Freq_gait).strip()

    Freq_EMG = (_read_param("ANALOG", "RATE"))
    
    if Freq_EMG is not None:
        result["Sample Rate EMG"] = str(Freq_EMG).strip()

    return result

def _normalize_side(ctx: str) -> str:
    """
    Map a CONTEXTS string from the C3D ANALYSIS block to one of
    "Right" / "Left" / "Bilateral".

    Handles case variations and common abbreviations:
        "Right", "RIGHT", "right", "R", "Rt", "rechts"  → "Right"
        "Left",  "LEFT",  "left",  "L", "Lt", "links"   → "Left"
        "General", "Both", "Bilateral", "Average", "",
        anything else                                    → "Bilateral"

    Bilateral parameters (Cadence, Walking Speed, Double Support, …) are
    stored once and copied into both Right/Left columns by the writer.
    """
    if not ctx:
        return "Bilateral"
    s = ctx.strip().lower()
    if s in ("right", "r", "rt", "rechts"):
        return "Right"
    if s in ("left", "l", "lt", "links"):
        return "Left"
    return "Bilateral"

def read_spatiotemporal_c3d(c3d: dict) -> dict[str, dict]:
    """
    Read ANALYSIS block using CONTEXTS to separate Right/Left/Bilateral.
    Returns {"Right":    {name: (value, unit)},
             "Left":     {name: (value, unit)},
             "Bilateral":{name: (value, unit)}}.

    Bilateral entries are parameters whose CONTEXTS slot is "General",
    empty, or otherwise non-sided (e.g. Cadence, Walking Speed). These
    are NOT dropped — the Spatiotemporal-sheet writer places them in
    both Right and Left columns for display.
    """
    result: dict[str, dict] = {"Right": {}, "Left": {}, "Bilateral": {}}
    try:
        ana      = c3d["parameters"]["ANALYSIS"]
        names    = [str(x).strip() for x in ana.get("NAMES",    {}).get("value", [])]
        contexts = [str(x).strip() for x in ana.get("CONTEXTS", {}).get("value", [])]
        values   = list(ana.get("VALUES", {}).get("value", []))
        units    = [str(x).strip() for x in ana.get("UNITS",    {}).get("value", [])]
        n = min(len(names), len(contexts), len(values))
        for i in range(n):
            name = names[i].strip()
            if not name:
                continue
            side = _normalize_side(contexts[i])
            v    = _safe_float(values[i])
            unit = units[i] if i < len(units) else ""
            if v is not None:
                result[side][name] = (v, unit)
    except Exception:
        pass
    return result

def read_clinical_c3d(c3d: dict) -> list[dict]:
    results: list[dict] = []
    try:
        proc = c3d["parameters"].get("PROCESSING", {})
    except Exception:
        return results
    for key in proc:
        if not any(kw in key.lower() for kw in _CLINICAL_KW):
            continue
        try:
            node = proc[key]
            v = None
            for leaf in ("value", "DATA"):
                if leaf in node:
                    v = node[leaf]; break
            if v is None: continue
            if isinstance(v, (list, tuple)): v = v[0] if v else None
            if np is not None and isinstance(v, np.ndarray):
                v = v.flat[0] if v.size > 0 else None
            if v is None: continue
            kl = key.lower()
            if "right" in kl or kl.startswith("r_") or "_r_" in kl: side = "Right"
            elif "left" in kl or kl.startswith("l_") or "_l_" in kl: side = "Left"
            else: side = "Both / Unspecified"
            results.append({"parameter": key, "value": v, "side": side})
        except Exception:
            pass
    return sorted(results, key=lambda x: x["parameter"])

#  PDF FALLBACK

def _pdf_fallback(c3d_path: Path) -> dict:
    """
    Look for a clinical PDF in the same folder as the C3D file.
    Returns a dict with keys: demo, st_text, clinical_text, raw_text
    All are plain strings/dicts for display — not structured like C3D blocks.
    Returns empty dict if PDF parser not available or no PDF found.
    """
    if not _PDF_PARSER_OK:
        return {}
    try:
        folder = c3d_path.parent
        pdf    = find_report_pdf(folder)
        if pdf is None:
            return {}
        raw_text = extract_text_from_pdf(pdf)
        fields   = parse_pdf_fields(raw_text)
        return {
            "pdf_path":  str(pdf),
            "fields":    fields,       # parsed structured fields
            "raw_text":  raw_text,     # full text for clinical fallback search
        }
    except Exception as exc:
        print(f"[_pdf_fallback] {exc}")
        return {}

def _demo_from_pdf(fb: dict) -> dict[str, object]:
    if not fb: return {}
    f = fb.get("fields", {})
    result: dict[str, object] = {}
    mapping = [
        ("last_name",   "Last Name"),
        ("first_name",  "First Name"),
        ("birth_date",  "Birth Date"),
        ("gender",      "Gender"),
        ("exam_date",   "Exam Date"),
        ("ganglabor_id","Gait Lab ID"),
        ("diagnosis",   "Diagnosis"),
        ("model",       "Model"),
    ]
    for key, label in mapping:
        v = f.get(key)
        if v:
            result[label] = v
    return result

def _st_from_pdf(fb: dict) -> dict[str, dict]:
    """
    Build ST dict from PDF parsed fields (measurements / conditions).
    These come out as free-text, not numeric, so we display them as-is.
    """
    if not fb: return {}
    f = fb.get("fields", {})
    result: dict[str, dict] = {"Right": {}, "Left": {}}
    meas = f.get("measurements", "")
    if meas:
        result["Right"]["Measurements"] = (meas, "")
        result["Left"]["Measurements"]  = (meas, "")
    cond_l = f.get("condition_left", "")
    cond_r = f.get("condition_right", "")
    if cond_l: result["Left"]["Condition"]  = (cond_l, "")
    if cond_r: result["Right"]["Condition"] = (cond_r, "")
    return result

def _clinical_from_pdf(fb: dict) -> list[dict]:
    if not fb: return []
    raw  = fb.get("raw_text", "")
    found: list[dict] = []
    for line in raw.splitlines():
        ll = line.lower()
        if any(kw in ll for kw in _CLINICAL_KW):
            line = line.strip()
            if len(line) < 3 or len(line) > 300:
                continue
            if "right" in ll or " r " in ll: side = "Right"
            elif "left" in ll  or " l " in ll: side = "Left"
            else: side = "Both / Unspecified"
            found.append({"parameter": "(PDF)", "value": line, "side": side})
    return found

#  EXCEL STYLE HELPERS

_COL_DARK   = "1A3A5C"
_COL_GREEN  = "0E5E35"
_COL_PURPLE = "4A235A"
_COL_RED    = "7B241C"
_COL_TEAL   = "0E6655"
_COL_ALT    = "EEF3FB"

def _hdr(ws, row: int, col: int, value: str, color: str = _COL_DARK) -> None:
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=9)
    c.fill      = PatternFill("solid", start_color=color)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    _brd(c)

def _dat(ws, row: int, col: int, value, alt: bool = False,
         fmt: str | None = None) -> None:
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Calibri", size=9)
    if alt:
        c.fill = PatternFill("solid", start_color=_COL_ALT)
    c.alignment = Alignment(vertical="center")
    if fmt:
        c.number_format = fmt
    _brd(c)

def _brd(c) -> None:
    bs = Side(style="thin", color="CCCCCC")
    c.border = Border(left=bs, right=bs, top=bs, bottom=bs)

def _section(ws, row: int, col: int, ncols: int, text: str,
             color: str = _COL_DARK) -> None:
    ws.merge_cells(start_row=row, start_column=col,
                   end_row=row,   end_column=col + ncols - 1)
    c = ws.cell(row=row, column=col, value=text)
    c.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    c.fill      = PatternFill("solid", start_color=color)
    c.alignment = Alignment(horizontal="left", vertical="center")

def _auto_width(ws, col: int, header: str, max_w: int = 28) -> None:
    ltr = get_column_letter(col)
    ws.column_dimensions[ltr].width = min(max(len(header) + 2, 10), max_w)

#  SHEET WRITERS

def _write_info_sheet(wb, c3d_path: Path, c3d: dict, pdf_path: str | None) -> None:
    ws = wb.create_sheet("Info")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 50
    _section(ws, 1, 1, 2, f"C3D File Info  —  {c3d_path.name}")

    fr_pt  = _point_rate(c3d)
    fr_an  = _analog_rate(c3d)
    try:
        n_fr   = c3d["data"]["points"].shape[2]
        n_pt   = c3d["data"]["points"].shape[1]
    except Exception:
        n_fr = n_pt = 0
    try:
        raw = c3d["data"]["analogs"]
        if raw.ndim == 3: raw = raw[0]
        n_an  = raw.shape[0]
        n_smp = raw.shape[1]
    except Exception:
        n_an = n_smp = 0

    # C3D header values (mirrors what tools like the C3D viewer show)
    # Note on indexing: ezc3d returns POINT.FIRST_FRAME / LAST_FRAME 0-indexed
    # (frame 1 in a 1-indexed C3D viewer = 0 in ezc3d). For DISPLAY we add +1
    # so the values match the C3D-viewer convention shown in Vicon/Nexus etc.
    # The Data Start Offset stays as raw_first_frame / rate — this IS the
    # correct elapsed-time-since-trial-start of the first recorded sample
    # (e.g. for raw=16, rate=150 → 16/150 = 0.1067 s, since frame "1" is at
    # t=0 and frame "17" is at t=16/150).
    raw_first  = _first_frame(c3d)
    raw_last   = _last_frame(c3d)
    first_frame = raw_first + 1                 # 1-indexed for display
    last_frame  = (raw_last + 1) if raw_last > 0 \
                  else (first_frame + n_fr - 1 if n_fr else first_frame)
    data_off_s  = _data_start_offset_s(c3d)     # = raw_first / fr_pt

    # Optional extras pulled directly from the header / parameters
    scale_factor = _get_scalar(c3d, "parameters", "POINT", "SCALE", "value")
    max_gap      = _get_scalar(c3d, "parameters", "POINT", "MAX_INTERPOLATION_GAP", "value")
    if max_gap is None:
        # ezc3d sometimes exposes it on the header
        try:
            max_gap = c3d["header"]["points"].get("nb_max_interpolation_gap")
        except Exception:
            max_gap = None

    rows = [
        ("C3D File",             str(c3d_path)),
        ("PDF Report",           pdf_path or "—  (not found)"),
        ("Point Sample Rate",    f"{fr_pt:.2f} Hz"),
        ("Analog Sample Rate",   f"{fr_an:.2f} Hz"),
        ("Number of Markers",    n_pt),
        ("Analog Channels",      n_an),
        ("First Frame",          first_frame),
        ("Last Frame",           last_frame),
        ("Point Frames",         n_fr),
        ("Analog Samples",       n_smp),
        ("Duration",             f"{n_fr / fr_pt:.3f} s"
                                 if fr_pt > 0 and n_fr > 0 else "—"),
        ("Data Start Offset (s)", round(data_off_s, 6)),
        ("Point Unit",           _point_unit(c3d) or "—"),
        ("Scale Factor",         scale_factor if scale_factor is not None else "—"),
        ("Maximum Interpolation Gap",
                                 max_gap if max_gap is not None else "—"),
    ]
    for i, (k, v) in enumerate(rows, 2):
        alt = i % 2 == 0
        _dat(ws, i, 1, k, alt)
        _dat(ws, i, 2, v if isinstance(v, (int, float)) else str(v), alt)

def _write_kv_sheet(wb, sheet_name: str, title: str, color: str,
                    items: list[tuple[str, object]],
                    col_headers: tuple[str, ...] = ("Parameter", "Value"),
                    source_note: str = "") -> None:
    """Generic key-value sheet."""
    ws = wb.create_sheet(sheet_name)
    nc = len(col_headers)
    _section(ws, 1, 1, nc,
             f"{title}" + (f"  [{source_note}]" if source_note else ""),
             color)
    for ci, h in enumerate(col_headers, 1):
        _hdr(ws, 2, ci, h, color)
    ws.column_dimensions["A"].width = 32
    for ci in range(2, nc + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 24
    ws.freeze_panes = "A3"

    if not items:
        ws.cell(row=3, column=1, value="⚠  No data found.")
        return
    for i, row_vals in enumerate(items, 3):
        alt = i % 2 == 0
        for ci, v in enumerate(row_vals, 1):
            _dat(ws, i, ci, v, alt)

def _write_demographics(wb, demo: dict[str, object], source: str) -> None:
    items = list(demo.items())
    _write_kv_sheet(wb, "Demographics", "Demographics", _COL_DARK,
                    items, ("Field", "Value"), source)

def _write_spatiotemporal(wb, st: dict[str, dict], source: str) -> None:
    ws = wb.create_sheet("Spatiotemporal")
    color = _COL_DARK
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    _section(ws, 1, 1, 4,
             f"Spatiotemporal Parameters" +
             (f"  [{source}]" if source else ""),
             color)
    for ci, h in enumerate(("Parameter", "Unit", "Right", "Left"), 1):
        _hdr(ws, 2, ci, h, color)
    ws.freeze_panes = "A3"

    # Merge bilateral entries (Cadence, Walking Speed, …) into both
    # Right and Left columns. Sided entries override bilateral ones
    # of the same name (e.g. an explicit "Right Cadence" wins over
    # a bilateral "Cadence" if both exist).
    bilateral = st.get("Bilateral", {})
    right = dict(bilateral); right.update(st.get("Right", {}))
    left  = dict(bilateral); left.update(st.get("Left", {}))

    seen: set[str] = set()
    params: list[str] = []
    for side_dict in (right, left):
        for k in side_dict:
            if k not in seen:
                params.append(k); seen.add(k)

    if not params:
        ws.cell(row=3, column=1, value="⚠  No spatiotemporal data found.")
        return

    for ri, param in enumerate(params, 3):
        alt = ri % 2 == 0
        r_entry = right.get(param)
        l_entry = left.get(param)
        r_val, r_unit = (r_entry if isinstance(r_entry, tuple) else (r_entry, "")) \
                         if r_entry else (None, "")
        l_val, l_unit = (l_entry if isinstance(l_entry, tuple) else (l_entry, "")) \
                         if l_entry else (None, "")
        # Both sides should report the same unit; pick whichever is non-empty.
        unit = r_unit or l_unit
        _dat(ws, ri, 1, param, alt)
        _dat(ws, ri, 2, unit or "", alt)
        _dat(ws, ri, 3, r_val, alt, fmt="0.00##")
        _dat(ws, ri, 4, l_val, alt, fmt="0.00##")

def _write_events(wb, events: list[dict]) -> None:
    ws = wb.create_sheet("Events")
    _section(ws, 1, 1, 4, "Gait Events  (from EVENT block)")
    for ci, h in enumerate(("No.", "Context", "Label", "Time (s)"), 1):
        _hdr(ws, 2, ci, h)
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 14
    ws.freeze_panes = "A3"
    if not events:
        ws.cell(row=3, column=1, value="⚠  No events found.")
        return
    for ri, ev in enumerate(events, 3):
        alt = ri % 2 == 0
        _dat(ws, ri, 1, ri - 2, alt)
        _dat(ws, ri, 2, ev["context"], alt)
        _dat(ws, ri, 3, ev["label"],   alt)
        _dat(ws, ri, 4, round(ev["time_s"], 5), alt, fmt="0.00000")

def _write_clinical(wb, items: list[dict], source: str) -> None:
    rows = [(it["parameter"], it["side"], it["value"]) for it in items]
    _write_kv_sheet(wb, "Clinical", "Clinical Data", _COL_PURPLE,
                    rows, ("Parameter", "Side", "Value"), source)

def _write_channel_sheet(
    wb,
    sheet_name: str,
    title:      str,
    color:      str,
    headers:    list[str],
    columns:    list[list],
    time_col:   bool = True, 
    rate:       float = 1.0,
    is_analog:  bool = False,
    max_rows:   int = 0, 
) -> None:
    """
    Generic wide-table sheet: rows = samples/frames, cols = channels.
    First col: Frame# (point data) or Time(s) (analog data).
    """
    ws = wb.create_sheet(sheet_name)
    n_cols = len(headers)

    if not headers:
        _section(ws, 1, 1, 2, f"⚠  {title}  — no channels found", color)
        return

    n_rows = len(columns[0]) if columns else 0
    if max_rows > 0 and n_rows > max_rows:
        note = f"  (first {max_rows:,} of {n_rows:,} samples shown)"
        n_rows = max_rows
    else:
        note = f"  ({n_rows:,} {'samples' if is_analog else 'frames'})"

    _section(ws, 1, 1, n_cols + 1, f"{title}{note}", color)

    # Row 2 — column headers
    time_hdr = "Time (s)" if is_analog else "Frame"
    _hdr(ws, 2, 1, time_hdr, color)
    ws.column_dimensions["A"].width = 12

    for ci, hdr in enumerate(headers, 2):
        _hdr(ws, 2, ci, hdr, color)
        # width heuristic
        w = min(max(len(hdr.split("(")[0].strip()) + 2, 12), 32)
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.row_dimensions[2].height = 50
    ws.freeze_panes = "B3"

    # Data rows
    num_fmt = "0.000000" if is_analog else "0.0000"
    for fi in range(n_rows):
        row = fi + 3
        alt = fi % 2 == 0
        t_val = round(fi / rate, 6) if is_analog else fi
        _dat(ws, row, 1, t_val, alt, fmt="0.000000" if is_analog else "0")
        for ci, col in enumerate(columns, 2):
            v = col[fi] if fi < len(col) else None
            _dat(ws, row, ci, v, alt, fmt=num_fmt)

#  PUBLIC ENTRY POINT

_MAX_ANALOG_ROWS = 50_000

def create_c3d_excel(
    c3d_path:  Path,
    out_path:  Path,
    opts:      dict | None = None,
    status_cb = None,
) -> None:
    """
    Read *c3d_path* and write a structured Excel workbook to *out_path*.

    opts keys (all default True):
        "info"       — Info sheet
        "demo"       — Demographics  (C3D first, PDF fallback)
        "st"         — Spatiotemporal (C3D first, PDF fallback)
        "clinical"   — Clinical data  (C3D first, PDF fallback)
        "events"     — Events sheet
        "angles"     — Angles sheet
        "moments"    — Moments sheet
        "powers"     — Powers sheet
        "forces"     — Forces (point-based) sheet
        "grf"        — GRF analog sheet
        "emg"        — EMG analog sheet
        "analogs_other" — remaining analogs
    """
    if np       is None: raise RuntimeError("numpy not installed   — pip install numpy")
    if ezc3d    is None: raise RuntimeError("ezc3d not installed    — pip install ezc3d")
    if openpyxl is None: raise RuntimeError("openpyxl not installed — pip install openpyxl")

    if opts is None:
        opts = {k: True for k in (
            "info", "demo", "st", "clinical", "events",
            "angles", "moments", "powers", "forces",
            "grf", "emg", "analogs_other")}

    def upd(msg: str) -> None:
        if status_cb: status_cb(msg)

    upd(f"Loading {c3d_path.name} …")
    c3d_obj  = ezc3d.c3d(str(c3d_path))
    fr_pt    = _point_rate(c3d_obj)
    fr_an    = _analog_rate(c3d_obj)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # PDF fallback (read once, reuse)
    pdf_fb = {}
    if _PDF_PARSER_OK:
        upd("Checking for paired PDF report …")
        pdf_fb = _pdf_fallback(c3d_path)
        if pdf_fb:
            upd(f"  PDF found: {Path(pdf_fb['pdf_path']).name}")
        else:
            upd("  No paired PDF found — C3D-only mode.")

    # 1. Info sheet
    upd("Writing file info …")
    _write_info_sheet(wb, c3d_path, c3d_obj,
                      pdf_fb.get("pdf_path") if pdf_fb else None)

    # 2. Demographics (C3D first → PDF fallback)
    if opts.get("demo", True):
        upd("Reading demographics (C3D PROCESSING block) …")
        demo = read_demographics_c3d(c3d_obj)
        source = "C3D PROCESSING block"
        if len(demo) < 2 and pdf_fb:
            upd("  C3D demographics sparse — trying PDF fallback …")
            pdf_demo = _demo_from_pdf(pdf_fb)
            if pdf_demo:
                # Merge: C3D wins where it has data, PDF fills gaps
                for k, v in pdf_demo.items():
                    if k not in demo:
                        demo[k] = v
                source = "C3D + PDF fallback"
        upd(f"  Demographics: {len(demo)} field(s)  [{source}]")
        _write_demographics(wb, demo, source)

    # 3. Spatiotemporal (C3D ANALYSIS block → PDF fallback)
    if opts.get("st", True):
        upd("Reading spatiotemporal (C3D ANALYSIS block) …")
        st = read_spatiotemporal_c3d(c3d_obj)
        n_st = sum(len(v) for v in st.values())
        source = "C3D ANALYSIS block"
        if n_st == 0 and pdf_fb:
            upd("  No C3D spatiotemporal — trying PDF fallback …")
            st = _st_from_pdf(pdf_fb)
            n_st = sum(len(v) for v in st.values())
            source = "PDF fallback"
        upd(f"  Spatiotemporal: {n_st} parameter(s)  [{source}]")
        _write_spatiotemporal(wb, st, source)

    # 4. Clinical (C3D PROCESSING keywords → PDF fallback)
    if opts.get("clinical", True):
        upd("Reading clinical data (C3D PROCESSING keywords) …")
        clinical = read_clinical_c3d(c3d_obj)
        source = "C3D PROCESSING block"
        if not clinical and pdf_fb:
            upd("  No C3D clinical data — trying PDF fallback …")
            clinical = _clinical_from_pdf(pdf_fb)
            source = "PDF text scan"
        upd(f"  Clinical: {len(clinical)} item(s)  [{source}]")
        _write_clinical(wb, clinical, source)

    # 5. Events
    if opts.get("events", True):
        upd("Reading events …")
        events = read_events(c3d_obj)
        upd(f"  Events: {len(events)}")
        _write_events(wb, events)

    # 6–9. POINT channels by type
    point_sheets = [
        ("angles",  "Angles",  _COL_TEAL,  _is_angle),
        ("moments", "Moments", _COL_GREEN, _is_moment),
        ("powers",  "Powers",  _COL_RED,   _is_power),
        ("forces",  "Forces",  "5D6D7E",   _is_force_pt),
    ]
    for opt_key, title, color, clf in point_sheets:
        if opts.get(opt_key, True):
            upd(f"Extracting {title} …")
            hdrs, cols = _extract_points_by_type(c3d_obj, clf)
            upd(f"  {title}: {len(hdrs)} columns")
            _write_channel_sheet(wb, title, title, color,
                                 hdrs, cols,
                                 rate=fr_pt, is_analog=False)

    # 10. GRF analogs
    if opts.get("grf", True):
        upd("Extracting GRF channels …")
        hdrs, cols = _extract_analogs_by_type(c3d_obj, _is_grf)
        upd(f"  GRF: {len(hdrs)} channels, "
            f"{len(cols[0]) if cols else 0:,} samples")
        _write_channel_sheet(wb, "GRF", "Ground Reaction Forces & Moments",
                             "884EA0", hdrs, cols,
                             rate=fr_an, is_analog=True,
                             max_rows=_MAX_ANALOG_ROWS)

    # 11. EMG analogs
    if opts.get("emg", True):
        upd("Extracting EMG channels …")
        hdrs, cols = _extract_analogs_by_type(c3d_obj, _is_emg)
        upd(f"  EMG: {len(hdrs)} channels, "
            f"{len(cols[0]) if cols else 0:,} samples")
        _write_channel_sheet(wb, "EMG", "EMG — All Channels", _COL_DARK,
                             hdrs, cols,
                             rate=fr_an, is_analog=True,
                             max_rows=_MAX_ANALOG_ROWS)

    upd(f"Saving {out_path.name} …")
    wb.save(str(out_path))
    upd("✓  Done.")

#  C3D WORKER THREAD

class C3DWorker(QThread):
    progress = Signal(str, int)       # message, percent
    file_status = Signal(str, str, str)  # path, status_text, status_type
    log_msg = Signal(str)
    finished = Signal(int, list)      # success_count, errors

    def __init__(self, files, out_dir, opts):
        super().__init__()
        self.files = files
        self.out_dir = out_dir
        self.opts = opts

    def run(self):
        total = len(self.files)
        success = 0
        errors = []

        for i, c3d_path in enumerate(self.files):
            self.progress.emit(
                f"[{i+1}/{total}]  {c3d_path.name}",
                int(100 * i / total),
            )
            self.file_status.emit(str(c3d_path), "Running…", "running")
            out_path = self.out_dir / (c3d_path.stem + ".xlsx")

            def _cb(msg, p=c3d_path):
                self.log_msg.emit(f"    {msg}")

            try:
                self.log_msg.emit(f"\n{'─'*50}\n{c3d_path.name}\n{'─'*50}")
                create_c3d_excel(c3d_path, out_path, self.opts, status_cb=_cb)
                success += 1
                self.file_status.emit(str(c3d_path), "✓ Done", "ok")
                self.log_msg.emit(f"  ✓  {c3d_path.name}  →  {out_path.name}")
            except Exception as exc:
                import traceback
                tb = traceback.format_exc()
                errors.append((c3d_path.name, str(exc)))
                self.file_status.emit(str(c3d_path), "✗ Error", "err")
                self.log_msg.emit(f"  ✗  {c3d_path.name}:  {exc}\n{tb}")

        self.finished.emit(success, errors)

#  UI TAB (PySide6)

class C3DExtractorTab(QWidget):

    def __init__(self, parent=None):
        super().__init__(parent)
        self._files: list[Path] = []
        self._out_dir: Path | None = None
        self._worker = None
        self._build()

    def _build(self) -> None:
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        # Header
        from GaitSharing_ui import make_page_header, make_accent_btn, make_console_log
        layout.addWidget(make_page_header(
            "C3D Extractor",
            "Extract data to structured Excel files (one per C3D)"))

        content = QWidget()
        clay = QVBoxLayout(content)
        clay.setContentsMargins(16, 12, 16, 12)
        clay.setSpacing(10)

        # File list
        file_group = QGroupBox("C3D Files")
        fglay = QVBoxLayout(file_group)

        btn_row = QHBoxLayout()
        add_btn = make_accent_btn("+ Add C3D Files")
        add_btn.clicked.connect(self._add_files)
        btn_row.addWidget(add_btn)
        rm_btn = QPushButton("Remove Selected")
        rm_btn.clicked.connect(self._remove_sel)
        btn_row.addWidget(rm_btn)
        clr_btn = QPushButton("Clear All")
        clr_btn.clicked.connect(self._clear)
        btn_row.addWidget(clr_btn)
        btn_row.addStretch()
        fglay.addLayout(btn_row)

        self.file_table = QTableWidget()
        self.file_table.setColumnCount(3)
        self.file_table.setHorizontalHeaderLabels(["Filename", "Folder", "Status"])
        self.file_table.setAlternatingRowColors(True)
        self.file_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.file_table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.file_table.verticalHeader().setVisible(False)
        self.file_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.file_table.horizontalHeader().setStretchLastSection(True)
        self.file_table.setColumnWidth(0, 200)
        self.file_table.setColumnWidth(1, 380)
        self.file_table.setMaximumHeight(140)
        fglay.addWidget(self.file_table)

        self._file_count_lbl = QLabel("No files selected.")
        self._file_count_lbl.setStyleSheet(f"color: {PALETTE['text_muted']}; font-size: 12px;")
        fglay.addWidget(self._file_count_lbl)
        clay.addWidget(file_group)

        # Output folder
        out_group = QGroupBox("Output Folder (one .xlsx per C3D file)")
        oglay = QHBoxLayout(out_group)
        self._out_lbl = QLabel("No folder selected.")
        self._out_lbl.setStyleSheet(f"color: {PALETTE['text_muted']};")
        oglay.addWidget(self._out_lbl, stretch=1)
        brw_btn = QPushButton("Browse…")
        brw_btn.clicked.connect(self._browse_out)
        oglay.addWidget(brw_btn)
        clay.addWidget(out_group)

        # Sheet selection
        opts_group = QGroupBox("Output Sheets")
        opts_lay = QVBoxLayout(opts_group)

        self._opt_demo    = QCheckBox("👤 Demographics");  self._opt_demo.setChecked(True)
        self._opt_st      = QCheckBox("🏃 Spatiotemporal"); self._opt_st.setChecked(True)
        self._opt_clinical= QCheckBox("🩺 Clinical");       self._opt_clinical.setChecked(True)
        self._opt_events  = QCheckBox("⚡ Events");        self._opt_events.setChecked(True)
        self._opt_angles  = QCheckBox("📐 Angles (deg)");  self._opt_angles.setChecked(True)
        self._opt_moments = QCheckBox("🔄 Moments (Nmm)"); self._opt_moments.setChecked(True)
        self._opt_powers  = QCheckBox("⚡ Powers (W)");    self._opt_powers.setChecked(True)
        self._opt_forces  = QCheckBox("↗ Forces/pt (N)");  self._opt_forces.setChecked(True)
        self._opt_grf     = QCheckBox("🦶 GRF (N)");       self._opt_grf.setChecked(True)
        self._opt_emg     = QCheckBox("💪 EMG");            self._opt_emg.setChecked(True)

        r1 = QHBoxLayout()
        for cb in [self._opt_demo, self._opt_st, self._opt_clinical, self._opt_events, self._opt_angles]:
            r1.addWidget(cb)
        r1.addStretch()
        opts_lay.addLayout(r1)

        r2 = QHBoxLayout()
        for cb in [self._opt_moments, self._opt_powers, self._opt_forces, self._opt_grf, self._opt_emg]:
            r2.addWidget(cb)
        r2.addStretch()
        opts_lay.addLayout(r2)
        clay.addWidget(opts_group)

        # Run button
        self._run_btn = make_accent_btn("▶   Extract Data")
        self._run_btn.clicked.connect(self._run)
        clay.addWidget(self._run_btn, alignment=Qt.AlignLeft)

        # Missing deps warning
        missing = [n for lib, n in ((np, "numpy"), (ezc3d, "ezc3d"),
                                     (openpyxl, "openpyxl")) if lib is None]
        if missing:
            warn = QLabel("⚠  Missing: " + ", ".join(missing) +
                          "   →   pip install " + " ".join(missing))
            warn.setStyleSheet(f"color: {PALETTE['warning']}; font-weight: 700;")
            clay.addWidget(warn)

        # Progress
        prog_group = QGroupBox("Progress")
        pglay = QVBoxLayout(prog_group)
        self._status_lbl = QLabel("Ready.")
        pglay.addWidget(self._status_lbl)
        self._pbar = QProgressBar()
        self._pbar.setMaximum(100)
        pglay.addWidget(self._pbar)
        clay.addWidget(prog_group)

        # Log
        log_group = QGroupBox("Log")
        lglay = QVBoxLayout(log_group)
        self._log_widget = make_console_log()
        self._log_widget.setMinimumHeight(100)
        lglay.addWidget(self._log_widget)
        clay.addWidget(log_group, stretch=1)

        # Wrap in scroll area so it works on small screens (14")
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setWidget(content)
        scroll.setStyleSheet("QScrollArea { border: none; }")
        layout.addWidget(scroll)

    # Helpers

    def _log(self, msg: str):
        self._log_widget.append(msg)

    def _refresh_file_list(self):
        self.file_table.setRowCount(len(self._files))
        for i, p in enumerate(self._files):
            self.file_table.setItem(i, 0, QTableWidgetItem(p.name))
            self.file_table.setItem(i, 1, QTableWidgetItem(str(p.parent)))
            item = QTableWidgetItem("Pending")
            item.setForeground(QColor(PALETTE["text_muted"]))
            self.file_table.setItem(i, 2, item)

        n = len(self._files)
        self._file_count_lbl.setText(
            f"{n} file{'s' if n != 1 else ''} selected." if n
            else "No files selected.")

    def _set_file_status(self, path_str: str, text: str, status_type: str):
        for row in range(self.file_table.rowCount()):
            item0 = self.file_table.item(row, 0)
            item1 = self.file_table.item(row, 1)
            if item0 and item1:
                full = str(Path(item1.text()) / item0.text())
                if full == path_str:
                    status_item = QTableWidgetItem(text)
                    color_map = {
                        "running": PALETTE["warning"],
                        "ok": PALETTE["success"],
                        "err": "#C42B1C",
                    }
                    status_item.setForeground(QColor(color_map.get(status_type, PALETTE["text"])))
                    self.file_table.setItem(row, 2, status_item)
                    break

    def _add_files(self):
        files, _ = QFileDialog.getOpenFileNames(
            self, "Select C3D Files",
            filter="C3D motion capture (*.c3d);;All files (*.*)")
        for raw in files:
            p = Path(raw)
            if p not in self._files:
                self._files.append(p)
        self._refresh_file_list()

    def _remove_sel(self):
        rows = sorted(set(idx.row() for idx in self.file_table.selectedIndexes()), reverse=True)
        for row in rows:
            if row < len(self._files):
                self._files.pop(row)
        self._refresh_file_list()

    def _clear(self):
        self._files.clear()
        self._refresh_file_list()

    def _browse_out(self):
        d = QFileDialog.getExistingDirectory(self, "Select Output Folder")
        if d:
            self._out_dir = Path(d)
            self._out_lbl.setText(str(self._out_dir))

    def _run(self):
        if not self._files:
            QMessageBox.warning(self, "No Files", "Add at least one .c3d file.")
            return
        if not self._out_dir:
            QMessageBox.critical(self, "No Output Folder", "Select an output folder.")
            return
        for lib, name in ((np, "numpy"), (ezc3d, "ezc3d"), (openpyxl, "openpyxl")):
            if lib is None:
                QMessageBox.critical(self, "Missing library",
                                     f"{name} not installed — pip install {name}")
                return

        self._run_btn.setEnabled(False)
        self._pbar.setValue(0)

        opts = {
            "demo":     self._opt_demo.isChecked(),
            "st":       self._opt_st.isChecked(),
            "clinical": self._opt_clinical.isChecked(),
            "events":   self._opt_events.isChecked(),
            "angles":   self._opt_angles.isChecked(),
            "moments":  self._opt_moments.isChecked(),
            "powers":   self._opt_powers.isChecked(),
            "forces":   self._opt_forces.isChecked(),
            "grf":      self._opt_grf.isChecked(),
            "emg":      self._opt_emg.isChecked(),
        }

        self._worker = C3DWorker(self._files[:], self._out_dir, opts)
        self._worker.progress.connect(
            lambda msg, pct: (self._status_lbl.setText(msg), self._pbar.setValue(pct)))
        self._worker.file_status.connect(self._set_file_status)
        self._worker.log_msg.connect(self._log)
        self._worker.finished.connect(self._on_finished)
        self._worker.start()

    def _on_finished(self, success: int, errors: list):
        total = len(self._files)
        self._pbar.setValue(100)
        self._status_lbl.setText(
            f"Done — {success}/{total} files, {len(errors)} error(s).")
        self._run_btn.setEnabled(True)
        if errors:
            QMessageBox.warning(self, "Done",
                f"{success} OK, {len(errors)} failed.\n\n" +
                "\n".join(f"• {n}: {e}" for n, e in errors[:5]))
        else:
            QMessageBox.information(self, "Done",
                f"✓  {success} file(s) saved to:\n{self._out_dir}")
