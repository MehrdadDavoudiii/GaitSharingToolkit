from __future__ import annotations

import re
from pathlib import Path

import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

#  STYLING (consistent with stride module)

_FONT_HDR   = Font(name="Arial", size=10, bold=True, color="FFFFFF")
_FONT_DATA  = Font(name="Arial", size=10)
_FONT_TITLE = Font(name="Arial", size=11, bold=True, color="FFFFFF")
_ALIGN_C    = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_L    = Alignment(horizontal="left", vertical="center")
_BORDER     = Border(
    bottom=Side(style="thin", color="DFE6E9"),
    right=Side(style="thin", color="DFE6E9"),
)
_FILL_ALT   = PatternFill("solid", fgColor="F8FAF0")
_NUM_FMT    = "0.0000"

_COL = {
    "Angles":   "1ABC9C",
    "Moments":  "27AE60",
    "Powers":   "E74C3C",
    "Forces":   "5D6D7E",
    "summary":  "9CBE20",
    "features": "2980B9",
}

_PHASES = ["Whole", "Stance", "Swing"]
_FEATURES = ["Max", "Min", "Mean", "Range", "Max@", "Min@"]

#  CELL HELPERS

def _section(ws, row, c0, c1, text, color="2D3436"):
    fill = PatternFill("solid", fgColor=color)
    for ci in range(c0, c1 + 1):
        ws.cell(row=row, column=ci).fill = fill
    c = ws.cell(row=row, column=c0, value=text)
    c.font = _FONT_TITLE
    c.alignment = _ALIGN_L
    ws.row_dimensions[row].height = 28

def _hdr(ws, row, col, text, color="2D3436"):
    c = ws.cell(row=row, column=col, value=text)
    c.font = _FONT_HDR
    c.fill = PatternFill("solid", fgColor=color)
    c.alignment = _ALIGN_C
    c.border = _BORDER

def _dat(ws, row, col, value, alt=False, fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = _FONT_DATA
    c.alignment = _ALIGN_L
    c.border = _BORDER
    if alt:
        c.fill = _FILL_ALT
    if fmt:
        c.number_format = fmt

#  EXCEL READER

_NORM_PATTERN = re.compile(
    r"^S(\d+)_(R|L)_(Angles|Moments|Powers|Forces)_Norm$"
)

def _parse_summary(wb) -> dict[str, list[float]]:
    """
    Read Summary sheet and return stance % per stride per side.
    Returns {"Right": [60.1, 61.3, …], "Left": [59.8, 60.5, …]}
    Each value = stance phase as % of gait cycle.
    """
    result: dict[str, list[float]] = {"Right": [], "Left": []}
    if "Summary" not in wb.sheetnames:
        return result

    ws = wb["Summary"]
    rows = list(ws.iter_rows(values_only=True))

    current_side = None
    for row in rows:
        if row[0] is None:
            continue
        cell0 = str(row[0]).strip()

        # Detect side header: "Right  —  2 stride(s)"
        if "Right" in cell0 and "stride" in cell0:
            current_side = "Right"
            continue
        elif "Left" in cell0 and "stride" in cell0:
            current_side = "Left"
            continue

        # Data rows: "S1", "S2", … in column A, Stance % in column F
        if current_side and cell0.startswith("S") and cell0[1:].isdigit():
            stance_pct = row[5] if len(row) > 5 else None
            if isinstance(stance_pct, (int, float)):
                result[current_side].append(float(stance_pct))
            else:
                # Default 60% if foot-off not detected
                result[current_side].append(60.0)

    return result

def _read_norm_sheet(wb, name: str) -> tuple[list[str], np.ndarray]:
    """
    Read a normalised sheet (101 rows × N channels).
    Returns (headers, data_2d) where data_2d.shape = (101, n_channels).
    """
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        return [], np.array([])

    header_row = rows[1]
    headers = [str(h).strip() for h in header_row[1:] if h is not None]
    n_ch = len(headers)
    if n_ch == 0:
        return [], np.array([])

    data = []
    for row in rows[2:]:
        if row[0] is None:
            break
        vals = []
        for ci in range(n_ch):
            v = row[ci + 1] if ci + 1 < len(row) else None
            try:
                vals.append(float(v))
            except (TypeError, ValueError):
                vals.append(np.nan)
        data.append(vals)

    return headers, np.array(data, dtype=float)

# Healthy-reference .txt parser

def _norm_key(s: str) -> str:
    return re.sub(r"\s+", " ", str(s).strip().lower())

def _parse_healthy_txt(path) -> dict[str, dict[str, dict[str, dict[str, float]]]]:
    """
    Parse a feature-summary .txt (format from _write_llm_text) into:
        { side: { channel_key: { phase: { "Max":v, "Min":v, "Mean":v,
                                          "Range":v, "Max@":v, "Min@":v } } } }
    Channel keys are lower-cased / whitespace-collapsed for tolerant matching.
    """
    result: dict[str, dict] = {"Right": {}, "Left": {}}
    if not path:
        return result
    try:
        text = Path(path).read_text(encoding="utf-8", errors="replace")
    except Exception:
        return result

    cur_side: str | None = None
    cur_ch:   str | None = None
    feat_re = re.compile(
        r"(Max|Min|Mean|Range)\s*=\s*(-?\d+(?:\.\d+)?)"
        r"(?:\s+at\s+(-?\d+(?:\.\d+)?)\s*%GC)?",
        re.IGNORECASE,
    )
    for raw in text.splitlines():
        line = raw.rstrip()
        if not line:
            continue
        stripped = line.strip()
        if stripped.upper().startswith("RIGHT SIDE"):
            cur_side, cur_ch = "Right", None;  continue
        if stripped.upper().startswith("LEFT SIDE"):
            cur_side, cur_ch = "Left",  None;  continue
        if cur_side is None:
            continue
        leading = len(line) - len(line.lstrip(" "))
        if leading <= 2 and stripped.endswith(":") and "=" not in stripped:
            cur_ch = _norm_key(stripped.rstrip(":"))
            result[cur_side].setdefault(cur_ch, {})
            continue
        if cur_ch is None:
            continue
        phase = None
        for pn in ("Whole Cycle", "Stance Phase", "Swing Phase"):
            if stripped.startswith(pn):
                phase = pn; break
        if phase is None:
            continue
        feats: dict[str, float] = {}
        for m in feat_re.finditer(stripped):
            name = m.group(1).capitalize()
            try:    feats[name] = float(m.group(2))
            except: continue
            if m.group(3) is not None:
                try:    feats[f"{name}@"] = float(m.group(3))
                except: pass
        if feats:
            result[cur_side][cur_ch][phase] = feats
    return result

def _healthy_lookup(healthy: dict, side: str, ch_name: str, phase: str
                    ) -> dict[str, float]:
    """Tolerant lookup; returns {} if any level is missing."""
    if not healthy:
        return {}
    return (healthy.get(side, {})
                   .get(_norm_key(ch_name), {})
                   .get(phase, {}))

#  FEATURE COMPUTATION

def _compute_features(signal: np.ndarray, pct_gc: np.ndarray) -> dict:
    """
    Compute 6 features from a 1-D signal segment.
    signal: array of values
    pct_gc: matching array of %GC values (0–100)
    Returns dict with keys: Max, Min, Mean, Range, Max@, Min@
    """
    if len(signal) == 0 or np.all(np.isnan(signal)):
        return {f: np.nan for f in _FEATURES}

    valid = ~np.isnan(signal)
    sig = signal[valid]
    pct = pct_gc[valid]

    if len(sig) == 0:
        return {f: np.nan for f in _FEATURES}

    mx = float(np.max(sig))
    mn = float(np.min(sig))
    return {
        "Max":   mx,
        "Min":   mn,
        "Mean":  float(np.mean(sig)),
        "Range": mx - mn,
        "Max@":  float(pct[np.argmax(sig)]),
        "Min@":  float(pct[np.argmin(sig)]),
    }

def _extract_features_for_stride(
    headers: list[str],
    data: np.ndarray,
    stance_pct: float,
) -> dict[str, float]:
    """
    Extract all features for one normalised stride.
    data.shape = (101, n_channels), rows = 0–100 %GC.

    Returns flat dict: {"channel__phase__feature": value, …}
    e.g. "RHipAngles_X__Whole__Max": 42.3
    """
    n_rows, n_ch = data.shape
    pct_gc = np.linspace(0, 100, n_rows)

    # Phase boundaries (as indices into 0–100)
    stance_idx = int(round(stance_pct))
    stance_idx = max(1, min(stance_idx, n_rows - 1))

    phases = {
        "Whole":  (0, n_rows),
        "Stance": (0, stance_idx),
        "Swing":  (stance_idx, n_rows),
    }

    features = {}
    for ci, hdr in enumerate(headers):
        # Shorten header for column name: "RHipAngles  [X / Sagittal] (deg)"
        # → "RHipAngles_X"
        short = _shorten_header(hdr)

        for phase_name, (i0, i1) in phases.items():
            seg = data[i0:i1, ci]
            pct = pct_gc[i0:i1]
            feats = _compute_features(seg, pct)
            for feat_name, val in feats.items():
                key = f"{short}__{phase_name}__{feat_name}"
                features[key] = val

    return features

def _shorten_header(hdr: str) -> str:
    """
    Shorten a full channel header to a compact ML-friendly name.
    "RHipAngles  [X / Sagittal] (deg)" → "RHipAngles_X"
    "LKneeMoment  [Z / Transverse] (Nmm)" → "LKneeMoment_Z"
    "LAnklePower  [X / Sagittal] (W)" → "LAnklePower_X"
    """
    # Extract name before bracket
    name = hdr.split("[")[0].strip()
    # Extract plane letter (X, Y, Z) from bracket
    m = re.search(r"\[([XYZ])", hdr)
    plane = m.group(1) if m else ""
    short = f"{name}_{plane}" if plane else name
    # Clean whitespace
    return re.sub(r"\s+", "", short)

#  EXCEL WRITER

# Plane mapping: short letter → (clinical plane, motion description)
_PLANE_MAP = {
    "X": ("Sagittal",    "Flex/Ext"),
    "Y": ("Frontal",     "Ab/Adduction"),
    "Z": ("Transverse",  "Int/Ext Rotation"),
}

# Side mapping
_SIDE_MAP = {"R": "Right", "L": "Left"}

# Known segment names → human-readable
_SEGMENT_NAMES = {
    "HipAngles":     "Hip Angles",
    "KneeAngles":    "Knee Angles",
    "AnkleAngles":   "Ankle Angles",
    "PelvisAngles":  "Pelvis Angles",
    "FootProgress":  "Foot Progression",
    "HipMoment":     "Hip Moment",
    "KneeMoment":    "Knee Moment",
    "AnkleMoment":   "Ankle Moment",
    "HipPower":      "Hip Power",
    "KneePower":     "Knee Power",
    "AnklePower":    "Ankle Power",
    "HipForce":      "Hip Force",
    "KneeForce":     "Knee Force",
    "AnkleForce":    "Ankle Force",
    "TrunkAngles":   "Trunk Angles",
    "SpineAngles":   "Spine Angles",
    "ShoulderAngles": "Shoulder Angles",
    "ThoraxAngles":  "Thorax Angles",
    "NormalisedGRF": "Normalised GRF",
}

# Data type → unit
_TYPE_UNITS = {
    "Angles": "deg",
    "Moments": "Nmm/kg",
    "Powers": "W/kg",
    "Forces": "N/kg",
}

def _expand_feature_name(key: str) -> dict:
    """
    Parse a compact ML feature key into human-readable components.

    Input:  "RHipAngles_X__Whole__Max"
    Output: { side, segment, plane, phase, feature, short, human, joint,
              plane_code, data_type }
    """
    parts = key.split("__")
    if len(parts) != 3:
        return {"human": key, "phase": "", "feature": "", "short": key,
                "side": "", "segment": "", "plane": "",
                "joint": "", "plane_code": "", "data_type": ""}

    channel, phase, feat = parts

    side_code = channel[0] if channel and channel[0] in ("R", "L") else ""
    side = _SIDE_MAP.get(side_code, "")

    rest = channel[1:] if side_code else channel
    if "_" in rest:
        seg_raw, plane_code = rest.rsplit("_", 1)
    else:
        seg_raw, plane_code = rest, ""

    segment = _SEGMENT_NAMES.get(seg_raw, seg_raw)
    plane_info = _PLANE_MAP.get(plane_code, (plane_code, ""))
    plane_str = f"{plane_info[0]} ({plane_info[1]})" if plane_info[1] else plane_info[0]

    phase_map = {"Whole": "Whole Cycle", "Stance": "Stance Phase", "Swing": "Swing Phase"}
    phase_human = phase_map.get(phase, phase)

    feat_map = {
        "Max": "Maximum", "Min": "Minimum", "Mean": "Mean",
        "Range": "Range", "Max@": "Max at (%GC)", "Min@": "Min at (%GC)",
    }
    feat_human = feat_map.get(feat, feat)

    human = f"{side} {segment}".strip()
    if plane_str:
        human += f" – {plane_str}"

    joint, data_type = _split_segment(seg_raw)

    return {
        "side": side, "segment": segment, "plane": plane_str,
        "phase": phase_human, "feature": feat_human,
        "short": channel, "human": human,
        "joint": joint, "plane_code": plane_code, "data_type": data_type,
    }

# joint / data_type extraction & filtering

_TYPE_SUFFIXES = [
    ("Angles",  "Angles"),
    ("Angle",   "Angles"),
    ("Moments", "Moments"),
    ("Moment",  "Moments"),
    ("Powers",  "Powers"),
    ("Power",   "Powers"),
    ("Forces",  "Forces"),
    ("Force",   "Forces"),
]

def _split_segment(seg_raw: str) -> tuple[str, str]:
    """Strip a trailing data-type token from a raw segment name."""
    if not seg_raw:
        return "", ""
    for suffix, dtype in _TYPE_SUFFIXES:
        if seg_raw.endswith(suffix) and len(seg_raw) > len(suffix):
            return seg_raw[: -len(suffix)], dtype
    return seg_raw, ""

_DEFAULT_JOINTS = [
    "Hip", "Knee", "Ankle",
    "Pelvis", "Spine", "Thorax", "Trunk",
    "Shoulder", "Elbow", "Wrist",
    "Head", "Neck",
    "FootProgress", "GroundReaction",
]

class FilterSpec:
    """
    Defines which (side, joint, plane, data_type) feature combinations are
    kept. ``healthy_path`` points to a healthy-reference .txt for comparison.
    """
    def __init__(self,
                 sides:       set[str] | None = None,
                 joints:      set[str] | None = None,
                 plane_codes: set[str] | None = None,
                 data_types:  set[str] | None = None,
                 phases:      set[str] | None = None,
                 custom_joints: list[str] | None = None,
                 healthy_path: str | Path | None = None):
        self.sides         = sides
        self.joints        = joints
        self.plane_codes   = plane_codes
        self.data_types    = data_types
        self.phases        = phases
        self.custom_joints = [c.strip().lower()
                              for c in (custom_joints or []) if c.strip()]
        self.healthy_path  = str(healthy_path) if healthy_path else None

    def accept(self, info: dict) -> bool:
        if self.sides is not None and info["side"] not in self.sides:
            return False
        if self.plane_codes is not None and info["plane_code"] not in self.plane_codes:
            return False
        if self.data_types is not None and info["data_type"] not in self.data_types:
            return False
        if self.phases is not None and info["phase"] not in self.phases:
            return False
        if self.joints is not None or self.custom_joints:
            j = info["joint"]
            in_list = (self.joints is not None and j in self.joints)
            in_custom = any(c in j.lower() for c in self.custom_joints)
            if not (in_list or in_custom):
                return False
        return True

    def describe(self) -> str:
        bits = []
        if self.sides is not None:
            bits.append("sides=" + "/".join(sorted(self.sides)))
        if self.joints is not None or self.custom_joints:
            jl = sorted(self.joints) if self.joints else []
            jl += [f"~{c}" for c in self.custom_joints]
            bits.append("joints=" + ",".join(jl))
        if self.plane_codes is not None:
            planes = [_PLANE_MAP.get(p, (p,))[0] for p in sorted(self.plane_codes)]
            bits.append("planes=" + "/".join(planes))
        if self.data_types is not None:
            bits.append("types=" + "/".join(sorted(self.data_types)))
        if self.healthy_path:
            bits.append(f"healthy={Path(self.healthy_path).name}")
        return "; ".join(bits) if bits else "all"

# stride-key helpers (handle both "S1_R" single-trial keys and
# "T1_S1_R" multi-trial keys produced by the averaged extractor)

def _stride_key_side(stride_key: str) -> str | None:
    """Return 'Right' / 'Left' / None from a stride key. Robust to T-prefix."""
    if stride_key.endswith("_R"):
        return "Right"
    if stride_key.endswith("_L"):
        return "Left"
    return None

def _stride_key_side_code(stride_key: str) -> str:
    """Return 'R' / 'L' / '' from a stride key."""
    if stride_key.endswith("_R"):
        return "R"
    if stride_key.endswith("_L"):
        return "L"
    return ""

def _stride_sort_key(k: str) -> tuple:
    """
    Sort key for stride identifiers — handles both 'S1_R' and 'T1_S1_R'.
    Sort order: trial idx (0 if absent), stride idx, side.
    """
    parts = k.split("_")
    trial = 0
    stride = 0
    side_code = ""
    for p in parts:
        if p.startswith("T") and p[1:].isdigit():
            trial = int(p[1:])
        elif p.startswith("S") and p[1:].isdigit():
            stride = int(p[1:])
        elif p in ("R", "L"):
            side_code = p
    return (trial, stride, side_code)

# stride-to-side averaging

def _aggregate_per_side(per_stride: dict[str, dict],
                         filter_spec: FilterSpec | None
                         ) -> tuple[dict[str, dict], dict[str, dict],
                                    dict[str, int]]:
    """
    Average feature values across strides of the same side, applying the
    optional filter so unwanted features are dropped before averaging.

    Returns:
        averaged:  {"Right": {feature_key: mean_value}, "Left": {...}}
        sd_vals:   {"Right": {feature_key: sd_value},   "Left": {...}}
        n_strides: {"Right": int, "Left": int} — count of source strides
    """
    by_side: dict[str, dict[str, list[float]]] = {"Right": {}, "Left": {}}
    n_strides: dict[str, int] = {"Right": 0, "Left": 0}

    for stride_key, feats in per_stride.items():
        side = _stride_key_side(stride_key)
        if side not in ("Right", "Left"):
            continue
        n_strides[side] += 1
        for key, val in feats.items():
            if filter_spec is not None:
                info = _expand_feature_name(key)
                if not filter_spec.accept(info):
                    continue
            if val is None:
                continue
            try:
                fv = float(val)
            except (TypeError, ValueError):
                continue
            if np.isnan(fv):
                continue
            by_side[side].setdefault(key, []).append(fv)

    averaged: dict[str, dict] = {"Right": {}, "Left": {}}
    sd_vals:  dict[str, dict] = {"Right": {}, "Left": {}}
    for side, kv in by_side.items():
        for key, vals in kv.items():
            if vals:
                averaged[side][key] = float(np.mean(vals))
                sd_vals[side][key]  = float(np.std(vals, ddof=1)) if len(vals) > 1 else 0.0
    return averaged, sd_vals, n_strides

def _count_per_feature(per_stride: dict[str, dict],
                        side_code: str,
                        feature_key: str) -> int:
    """Count strides on a side that contributed a non-NaN value for a feature."""
    n = 0
    for stride_key, feats in per_stride.items():
        if _stride_key_side_code(stride_key) != side_code:
            continue
        v = feats.get(feature_key)
        if v is None:
            continue
        try:
            fv = float(v)
        except (TypeError, ValueError):
            continue
        if not np.isnan(fv):
            n += 1
    return n

def _write_clinical_summary_sheet(wb,
                                   per_stride: dict[str, dict],
                                   filter_spec: FilterSpec | None,
                                   source: str,
                                   healthy: dict | None = None,
                                   title_prefix: str = "Clinical Feature Summary",
                                   ) -> tuple[int, dict[str, int]]:
    """
    Clinical_Summary: per-side averaged features with SD and optional
    healthy-reference comparison.
    """
    averaged, sd_vals, n_strides = _aggregate_per_side(per_stride, filter_spec)
    has_h = bool(healthy)

    ws = wb.create_sheet("Clinical_Summary")
    color = "2980B9"
    ws.sheet_properties.tabColor = color

    col_hdrs = ["Side", "Joint / Channel", "Plane", "Phase", "N strides",
                "Max", "Max SD", "Min", "Min SD", "Mean", "Mean SD",
                "Range", "Range SD", "Max @%GC", "Max@ SD",
                "Min @%GC", "Min@ SD"]
    if has_h:
        col_hdrs += ["Healthy Max", "Healthy Min", "Healthy Mean",
                     "Healthy Range", "Healthy Max @%GC", "Healthy Min @%GC"]
    n_cols = len(col_hdrs)

    sub = filter_spec.describe() if filter_spec is not None else "all"
    _section(ws, 1, 1, n_cols,
             f"{title_prefix}  —  averaged ± SD"
             f"   [filter: {sub}]   [{source}]", color)

    for ci, h in enumerate(col_hdrs, 1):
        _hdr(ws, 2, ci, h, color)

    widths = [10, 26, 22, 14, 7,
              10, 7, 10, 7, 10, 7, 10, 7, 9, 7, 9, 7]
    if has_h:
        widths += [10, 10, 10, 10, 10, 10]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 30
    ws.freeze_panes = "A3"

    _FEAT_ORDER = [
        ("Maximum",       False),
        ("Minimum",       False),
        ("Mean",          False),
        ("Range",         False),
        ("Max at (%GC)",  True),
        ("Min at (%GC)",  True),
    ]

    row = 3
    n_written = 0
    for side, side_code in (("Right", "R"), ("Left", "L")):
        feats_avg = averaged.get(side, {})
        feats_sd  = sd_vals.get(side, {})
        if not feats_avg:
            continue

        groups:    dict[tuple[str, str], dict[str, float]] = {}
        sd_groups: dict[tuple[str, str], dict[str, float]] = {}
        meta: dict[str, dict] = {}
        for key, val in feats_avg.items():
            info = _expand_feature_name(key)
            ch, phase, feat_h = info["human"], info["phase"], info["feature"]
            groups.setdefault((ch, phase), {})[feat_h] = val
            sd_v = feats_sd.get(key, 0.0)
            sd_groups.setdefault((ch, phase), {})[feat_h] = sd_v
            meta.setdefault(ch, {"plane": info["plane"],
                                 "first_key": key})

        ch_order = sorted(meta.keys())
        for ch in ch_order:
            plane = meta[ch]["plane"]
            for phase_name in ("Whole Cycle", "Stance Phase", "Swing Phase"):
                if (ch, phase_name) not in groups:
                    continue
                fv = groups[(ch, phase_name)]
                sv = sd_groups[(ch, phase_name)]
                first_key = meta[ch]["first_key"]
                n = _count_per_feature(per_stride, side_code, first_key)

                alt = row % 2 == 0
                _dat(ws, row, 1, side, alt)
                _dat(ws, row, 2, ch, alt)
                _dat(ws, row, 3, plane, alt)
                _dat(ws, row, 4, phase_name, alt)
                _dat(ws, row, 5, n, alt)

                ci = 6
                for fn, is_timing in _FEAT_ORDER:
                    v = fv.get(fn)
                    s = sv.get(fn, 0.0)
                    fmt = "0.0" if is_timing else _NUM_FMT
                    if v is not None and not np.isnan(v):
                        _dat(ws, row, ci, round(float(v), 4), alt, fmt)
                    else:
                        _dat(ws, row, ci, None, alt)
                    ci += 1
                    if s is not None and n > 1 and not np.isnan(s):
                        _dat(ws, row, ci, round(float(s), 4), alt, fmt)
                    else:
                        _dat(ws, row, ci, None, alt)
                    ci += 1

                if has_h:
                    hv = _healthy_lookup(healthy, side, ch, phase_name)
                    for hkey in ("Max", "Min", "Mean", "Range", "Max@", "Min@"):
                        v = hv.get(hkey)
                        fmt = "0.0" if "@" in hkey else _NUM_FMT
                        if v is not None:
                            _dat(ws, row, ci, round(float(v), 4), alt, fmt)
                        else:
                            _dat(ws, row, ci, None, alt)
                        ci += 1

                row += 1
                n_written += 1

    if n_written == 0:
        ws.cell(row=3, column=1,
                value="⚠  No features matched the current filter.")
    return n_written, n_strides

def _write_stride_details(wb,
                          per_stride: dict[str, dict],
                          filter_spec: FilterSpec | None,
                          source: str,
                          title_prefix: str = "Stride Details",
                          ) -> int:
    """
    Write a Stride_Details sheet: one row per (stride, channel, phase).
    Same filter as Clinical_Summary but NO averaging — raw per-stride values.
    """
    ws = wb.create_sheet("Stride_Details")
    color = "1ABC9C"
    ws.sheet_properties.tabColor = color

    col_hdrs = ["Stride", "Side", "Joint / Channel", "Plane", "Phase",
                "Max", "Min", "Mean", "Range", "Max @%GC", "Min @%GC"]
    n_cols = len(col_hdrs)

    sub = filter_spec.describe() if filter_spec is not None else "all"
    _section(ws, 1, 1, n_cols,
             f"{title_prefix}  —  per-stride values (before averaging)"
             f"   [filter: {sub}]   [{source}]", color)

    for ci, h in enumerate(col_hdrs, 1):
        _hdr(ws, 2, ci, h, color)

    widths = [10, 8, 26, 22, 14, 10, 10, 10, 10, 9, 9]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 30
    ws.freeze_panes = "A3"

    _FEAT_NAMES = ["Maximum", "Minimum", "Mean", "Range",
                   "Max at (%GC)", "Min at (%GC)"]

    row = 3
    n_written = 0
    for stride_key in sorted(per_stride.keys(), key=_stride_sort_key):
        feats = per_stride[stride_key]
        if not feats:
            continue
        side_code = _stride_key_side_code(stride_key)
        side = _SIDE_MAP.get(side_code, side_code)

        groups: dict[tuple[str, str], dict[str, float]] = {}
        meta: dict[str, dict] = {}
        for key, val in feats.items():
            info = _expand_feature_name(key)
            if filter_spec is not None and not filter_spec.accept(info):
                continue
            ch, phase, feat_h = info["human"], info["phase"], info["feature"]
            groups.setdefault((ch, phase), {})[feat_h] = val
            meta.setdefault(ch, {"plane": info["plane"]})

        ch_order = sorted(meta.keys())
        for ch in ch_order:
            plane = meta[ch]["plane"]
            for phase_name in ("Whole Cycle", "Stance Phase", "Swing Phase"):
                if (ch, phase_name) not in groups:
                    continue
                fv = groups[(ch, phase_name)]
                alt = row % 2 == 0
                _dat(ws, row, 1, stride_key, alt)
                _dat(ws, row, 2, side, alt)
                _dat(ws, row, 3, ch, alt)
                _dat(ws, row, 4, plane, alt)
                _dat(ws, row, 5, phase_name, alt)
                for ci, fn in enumerate(_FEAT_NAMES, 6):
                    v = fv.get(fn)
                    if v is not None and not np.isnan(v):
                        fmt = "0.0" if "@" in fn else _NUM_FMT
                        _dat(ws, row, ci, round(float(v), 4), alt, fmt)
                    else:
                        _dat(ws, row, ci, None, alt)
                row += 1
                n_written += 1

    if n_written == 0:
        ws.cell(row=3, column=1,
                value="⚠  No features matched the current filter.")
    return n_written

def _write_llm_text(output_path,
                     per_stride: dict[str, dict],
                     filter_spec: FilterSpec | None,
                     stance_data: dict[str, list[float]],
                     header_lines: list[str] | None = None):
    """
    Write a companion .txt file with averaged-per-side feature summary.
    Same filter as the Clinical_Summary sheet so both files agree.

    ``header_lines`` allow callers (e.g. multi-trial averager) to add extra
    context above the body — kept compatible with single-trial output.
    """
    txt_path = output_path.with_suffix(".txt")
    averaged, _sd, n_strides = _aggregate_per_side(per_stride, filter_spec)

    lines = []
    lines.append("=" * 70)
    if header_lines and header_lines[0]:
        lines.append(header_lines[0])
    else:
        lines.append("GAIT FEATURE SUMMARY  —  averaged across strides per side")
    lines.append("=" * 70)
    lines.append(f"Source: {output_path.stem}")
    if filter_spec is not None:
        lines.append(f"Filter: {filter_spec.describe()}")
    if header_lines and len(header_lines) > 1:
        for hl in header_lines[1:]:
            if hl:
                lines.append(hl)
    lines.append("")

    for side, side_code in (("Right", "R"), ("Left", "L")):
        feats_avg = averaged.get(side, {})
        if not feats_avg:
            continue

        stances = stance_data.get(side, [])
        if stances:
            stance_mean = float(np.mean(stances))
            stance_str = (f"  (Stance {stance_mean:.1f}% / "
                          f"Swing {100 - stance_mean:.1f}%)")
        else:
            stance_str = ""

        n = n_strides.get(side, 0)
        lines.append("-" * 70)
        lines.append(f"{side.upper()} SIDE  —  averaged over {n} stride(s)"
                     + stance_str)
        lines.append("-" * 70)

        channels: dict[str, dict[str, dict[str, float]]] = {}
        for key, val in feats_avg.items():
            info = _expand_feature_name(key)
            ch = info["human"]
            phase = info["phase"]
            feat = info["feature"]
            channels.setdefault(ch, {}).setdefault(phase, {})[feat] = val

        for ch_name in sorted(channels):
            phases = channels[ch_name]
            lines.append(f"\n  {ch_name}:")
            for phase_name in ("Whole Cycle", "Stance Phase", "Swing Phase"):
                if phase_name not in phases:
                    continue
                fv = phases[phase_name]
                mx     = fv.get("Maximum")
                mn     = fv.get("Minimum")
                mean   = fv.get("Mean")
                rng    = fv.get("Range")
                mx_at  = fv.get("Max at (%GC)")
                mn_at  = fv.get("Min at (%GC)")
                bits = []
                if mx is not None and not np.isnan(mx):
                    s = f"Max={mx:.1f}"
                    if mx_at is not None and not np.isnan(mx_at):
                        s += f" at {mx_at:.0f}%GC"
                    bits.append(s)
                if mn is not None and not np.isnan(mn):
                    s = f"Min={mn:.1f}"
                    if mn_at is not None and not np.isnan(mn_at):
                        s += f" at {mn_at:.0f}%GC"
                    bits.append(s)
                if mean is not None and not np.isnan(mean):
                    bits.append(f"Mean={mean:.1f}")
                if rng is not None and not np.isnan(rng):
                    bits.append(f"Range={rng:.1f}")
                if bits:
                    lines.append(f"    {phase_name:14s}  {', '.join(bits)}")

    lines.append("\n" + "=" * 70)
    lines.append("END OF FEATURE SUMMARY")
    lines.append("=" * 70)

    txt_path.write_text("\n".join(lines), encoding="utf-8")
    return txt_path

#  PER-STRIDE EXTRACTION HELPER
#  Pulled out of extract_gait_features so it can be reused by the
#  multi-trial averaged extractor.

def _extract_per_stride_from_excel(
    input_path: Path,
    status_cb=None,
) -> tuple[dict[str, dict], dict[str, list[float]], list[str]]:
    """
    Read one stride-analysis .xlsx and compute the per-stride feature dict.

    Returns:
        per_stride:  {"S1_R": {feature_key: value, …}, "S2_R": {…}, …}
        stance_data: {"Right": [60.1, 61.3, …], "Left": [...]}
        stride_keys: ordered list of stride keys (e.g. ["S1_R", "S1_L", …])
    """
    def upd(msg):
        if status_cb:
            status_cb(msg)

    upd(f"Loading {input_path.name} …")
    wb_in = openpyxl.load_workbook(str(input_path), data_only=True,
                                    read_only=True)

    stance_data = _parse_summary(wb_in)
    upd(f"  Stance %: Right={stance_data['Right']}, Left={stance_data['Left']}")

    norm_sheets = []
    for name in wb_in.sheetnames:
        m = _NORM_PATTERN.match(name)
        if m:
            stride_num = int(m.group(1))
            side_tag   = m.group(2)
            data_type  = m.group(3)
            side       = "Right" if side_tag == "R" else "Left"
            norm_sheets.append({
                "name":       name,
                "stride_num": stride_num,
                "side_tag":   side_tag,
                "side":       side,
                "data_type":  data_type,
            })

    norm_sheets.sort(key=lambda x: (x["stride_num"], x["side_tag"]))
    upd(f"  Found {len(norm_sheets)} normalised sheets")

    if not norm_sheets:
        wb_in.close()
        raise ValueError("No normalised sheets found (S*_*_*_Norm).")

    stride_keys: list[str] = []
    seen_keys: set[str] = set()
    for ns in norm_sheets:
        key = f"S{ns['stride_num']}_{ns['side_tag']}"
        if key not in seen_keys:
            stride_keys.append(key)
            seen_keys.add(key)

    per_stride: dict[str, dict] = {k: {} for k in stride_keys}

    for ns in norm_sheets:
        name = ns["name"]
        side = ns["side"]
        si   = ns["stride_num"] - 1
        key  = f"S{ns['stride_num']}_{ns['side_tag']}"

        upd(f"  Reading {name} …")
        headers, data = _read_norm_sheet(wb_in, name)
        if not headers or data.size == 0:
            upd(f"    Empty — skipping")
            continue

        stance_list = stance_data.get(side, [])
        stance_pct  = stance_list[si] if si < len(stance_list) else 60.0
        upd(f"    {len(headers)} channels, stance={stance_pct:.1f}%")

        features = _extract_features_for_stride(headers, data, stance_pct)
        per_stride[key].update(features)

    wb_in.close()
    return per_stride, stance_data, stride_keys

#  MAIN PUBLIC FUNCTION — single trial

def extract_gait_features(
    input_path:  str | Path,
    output_path: str | Path,
    filter_spec: FilterSpec | None = None,
    status_cb=None,
) -> dict:
    """
    Read one stride-analysis Excel and write a per-side averaged
    Clinical_Summary sheet plus a companion .txt for LLM consumption.
    """
    input_path  = Path(input_path)
    output_path = Path(output_path)

    def upd(msg):
        if status_cb:
            status_cb(msg)

    # 1. Read strides
    per_stride, stance_data, stride_keys = _extract_per_stride_from_excel(
        input_path, status_cb=status_cb)

    # 2. Parse healthy reference (if provided)
    healthy = None
    if filter_spec is not None and filter_spec.healthy_path:
        upd(f"  Loading healthy reference: "
            f"{Path(filter_spec.healthy_path).name}")
        healthy = _parse_healthy_txt(filter_spec.healthy_path)
        n_h = sum(len(c) for c in healthy.values())
        if n_h == 0:
            upd("  ⚠  Healthy file empty / unreadable — proceeding "
                "without comparison.")
            healthy = None
        else:
            upd(f"    parsed {n_h} channel entr"
                f"{'y' if n_h == 1 else 'ies'} from healthy reference")

    # 3. Write output
    upd(f"Writing Clinical_Summary  [filter: "
        f"{filter_spec.describe() if filter_spec else 'all'}] …")
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    n_rows, n_strides = _write_clinical_summary_sheet(
        wb_out, per_stride, filter_spec, input_path.name, healthy=healthy)

    upd("Writing Stride_Details …")
    n_detail = _write_stride_details(
        wb_out, per_stride, filter_spec, input_path.name)
    upd(f"  → {n_detail} detail rows")

    upd(f"Saving {output_path.name} …")
    wb_out.save(str(output_path))

    # 4. LLM text summary (no healthy — patient only)
    upd("Writing LLM text summary …")
    txt_path = _write_llm_text(output_path, per_stride,
                                filter_spec, stance_data)
    upd(f"  → {txt_path.name}")

    upd(f"✓ Done — {n_rows} summary rows + {n_detail} detail rows  "
        f"(Right: {n_strides['Right']} strides, "
        f"Left: {n_strides['Left']} strides"
        + (f", with healthy reference" if healthy else "")
        + ")")

    return {
        "n_rows":         n_rows,
        "n_detail_rows":  n_detail,
        "n_strides_R":    n_strides["Right"],
        "n_strides_L":    n_strides["Left"],
        "stride_labels":  stride_keys,
        "txt_path":       str(txt_path),
        "filter":         filter_spec.describe() if filter_spec else "all",
        "healthy_used":   healthy is not None,
    }

#  MAIN PUBLIC FUNCTION — multi-trial average

def extract_features_averaged(
    input_paths: list[str | Path],
    output_path: str | Path,
    filter_spec: FilterSpec | None = None,
    status_cb=None,
) -> dict:
    """
    Pool strides from multiple stride-analysis Excel files and write a single
    averaged Clinical_Summary sheet + companion .txt summarising the patient
    across all input trials.

    Strategy: every stride from every input file is treated as one big set,
    averaged per side. This gives more strides per side and tighter SDs than
    averaging trial-level means.
    """
    output_path = Path(output_path)

    def upd(msg):
        if status_cb:
            status_cb(msg)

    if not input_paths:
        raise ValueError("extract_features_averaged: input_paths is empty")

    # 1. Read each trial; merge per_stride dicts with T-prefixed keys
    merged: dict[str, dict] = {}
    merged_stance: dict[str, list[float]] = {"Right": [], "Left": []}
    trial_summaries: list[str] = []

    for trial_idx, raw in enumerate(input_paths, start=1):
        p = Path(raw)
        upd(f"\n── Trial {trial_idx}/{len(input_paths)}: {p.name} ──")
        try:
            per_stride, stance_data, stride_keys = \
                _extract_per_stride_from_excel(p, status_cb=status_cb)
        except Exception as exc:
            upd(f"  ⚠  Skipping {p.name}: {exc}")
            continue

        # T-prefix keys so they stay unique across trials
        nR = nL = 0
        for k, v in per_stride.items():
            new_key = f"T{trial_idx}_{k}"
            merged[new_key] = v
            sc = _stride_key_side_code(new_key)
            if sc == "R": nR += 1
            elif sc == "L": nL += 1

        for side in ("Right", "Left"):
            merged_stance[side].extend(stance_data.get(side, []))

        trial_summaries.append(f"T{trial_idx}: {p.name} (R:{nR}, L:{nL})")

    if not merged:
        raise ValueError(
            "No usable strides extracted from any of the input files.")

    # 2. Optional healthy reference
    healthy = None
    if filter_spec is not None and filter_spec.healthy_path:
        upd(f"\n  Loading healthy reference: "
            f"{Path(filter_spec.healthy_path).name}")
        healthy = _parse_healthy_txt(filter_spec.healthy_path)
        if sum(len(c) for c in healthy.values()) == 0:
            healthy = None
            upd("  ⚠  Healthy file empty — proceeding without comparison.")

    # 3. Write averaged Excel
    n_trials = len(trial_summaries)
    src_label = f"AVERAGE OF {n_trials} TRIAL(S)"
    upd(f"\nWriting averaged Clinical_Summary  [filter: "
        f"{filter_spec.describe() if filter_spec else 'all'}] …")
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    n_rows, n_strides = _write_clinical_summary_sheet(
        wb_out, merged, filter_spec, src_label, healthy=healthy,
        title_prefix=f"Multi-Trial Clinical Feature Summary ({n_trials} trials)")

    upd("Writing per-stride details (across all trials) …")
    n_detail = _write_stride_details(
        wb_out, merged, filter_spec, src_label,
        title_prefix=f"Stride Details across {n_trials} trials")
    upd(f"  → {n_detail} detail rows")

    # Add a small "Trials" sheet listing what went in
    ws_t = wb_out.create_sheet("Trials", 0)
    ws_t.sheet_properties.tabColor = "8E44AD"
    _section(ws_t, 1, 1, 3, f"Source trials  —  {n_trials} file(s) "
             f"({n_strides['Right']} R / {n_strides['Left']} L strides total)",
             "8E44AD")
    _hdr(ws_t, 2, 1, "Trial", "8E44AD")
    _hdr(ws_t, 2, 2, "File name", "8E44AD")
    _hdr(ws_t, 2, 3, "Stride counts", "8E44AD")
    ws_t.column_dimensions["A"].width = 8
    ws_t.column_dimensions["B"].width = 60
    ws_t.column_dimensions["C"].width = 22
    for i, summary in enumerate(trial_summaries, start=3):
        # summary is "T1: name.xlsx (R:n, L:n)"
        m = re.match(r"^(T\d+):\s+(.*?)\s+\((R:\d+,\s*L:\d+)\)\s*$", summary)
        if m:
            t, name, counts = m.group(1), m.group(2), m.group(3)
        else:
            t, name, counts = "", summary, ""
        alt = (i % 2 == 0)
        _dat(ws_t, i, 1, t, alt)
        _dat(ws_t, i, 2, name, alt)
        _dat(ws_t, i, 3, counts, alt)

    upd(f"Saving {output_path.name} …")
    wb_out.save(str(output_path))

    # 4. LLM text — multi-trial flavour
    upd("Writing multi-trial LLM text summary …")
    header_lines = [
        f"GAIT FEATURE SUMMARY  —  POOLED ACROSS {n_trials} TRIAL(S)",
        f"Trials pooled:",
    ]
    for s in trial_summaries:
        header_lines.append(f"  • {s}")
    txt_path = _write_llm_text(output_path, merged,
                                filter_spec, merged_stance,
                                header_lines=header_lines)
    upd(f"  → {txt_path.name}")

    upd(f"✓ Multi-trial average done — {n_rows} summary rows + "
        f"{n_detail} detail rows  "
        f"(Right: {n_strides['Right']} strides, "
        f"Left: {n_strides['Left']} strides, "
        f"from {n_trials} trial(s)"
        + (f", with healthy reference" if healthy else "")
        + ")")

    return {
        "n_trials":       n_trials,
        "n_rows":         n_rows,
        "n_detail_rows":  n_detail,
        "n_strides_R":    n_strides["Right"],
        "n_strides_L":    n_strides["Left"],
        "trials":         trial_summaries,
        "txt_path":       str(txt_path),
        "filter":         filter_spec.describe() if filter_spec else "all",
        "healthy_used":   healthy is not None,
    }

# Helper to derive a sensible name for the multi-trial output
def _suggest_average_stem(input_paths: list[Path]) -> str:
    """
    Build an output stem for the multi-trial average file.
    Uses the longest common prefix of input stems if non-trivial,
    otherwise the parent folder name, otherwise a generic fallback.
    """
    stems = [Path(p).stem for p in input_paths]
    if not stems:
        return "MULTI_TRIAL_AVERAGE_features"
    common = stems[0]
    for s in stems[1:]:
        i = 0
        while i < len(common) and i < len(s) and common[i] == s[i]:
            i += 1
        common = common[:i]
    common = common.rstrip(" _-.")
    if len(common) >= 3:
        return f"{common}_AVERAGE_features"
    parent = Path(input_paths[0]).parent.name
    if parent:
        return f"{parent}_AVERAGE_features"
    return "MULTI_TRIAL_AVERAGE_features"

#  PySide6 WORKER + TAB

try:
    from PySide6.QtCore import Qt, Signal, QThread, QSettings
    from PySide6.QtWidgets import (
        QWidget, QVBoxLayout, QHBoxLayout, QGridLayout, QGroupBox, QLabel,
        QPushButton, QProgressBar, QFileDialog, QMessageBox,
        QTableWidget, QTableWidgetItem, QAbstractItemView,
        QScrollArea, QDialog, QDialogButtonBox, QCheckBox, QLineEdit,
    )
    from PySide6.QtGui import QColor
    _HAS_PYSIDE = True
except ImportError:
    _HAS_PYSIDE = False

if _HAS_PYSIDE:

    from GaitSharing_config import PALETTE

    # Filter selection dialog
    class FilterSelectionDialog(QDialog):
        """
        Pop-up for selecting features + optional healthy reference comparison
        + multi-trial average toggle. Persisted via QSettings.
        """
        _SETTINGS_ORG = "GaitSharing"
        _SETTINGS_APP = "FeatureExtractor"

        def __init__(self, parent=None, multi_trial: bool = False):
            super().__init__(parent)
            self._multi_trial = multi_trial
            self.setWindowTitle("Select Features to Include")
            self.setModal(True)
            self.resize(680, 700)
            self._build()
            self._load_settings()

        def _build(self):
            layout = QVBoxLayout(self)
            layout.setSpacing(10)

            intro = QLabel(
                "Choose which features the Clinical_Summary sheet should "
                "contain. Values are averaged across strides per side."
            )
            intro.setWordWrap(True)
            intro.setStyleSheet(
                f"color: {PALETTE['text_muted']}; font-size: 12px;")
            layout.addWidget(intro)

            top = QHBoxLayout()
            layout.addLayout(top)

            left_col = QVBoxLayout()
            top.addLayout(left_col, stretch=1)

            # Sides
            side_grp = QGroupBox("Side")
            side_lay = QVBoxLayout(side_grp)
            self._cb_sides = {
                "Right": QCheckBox("Right"),
                "Left":  QCheckBox("Left"),
            }
            for cb in self._cb_sides.values():
                cb.setChecked(True)
                side_lay.addWidget(cb)
            left_col.addWidget(side_grp)

            # Planes
            plane_grp = QGroupBox("Plane")
            plane_lay = QVBoxLayout(plane_grp)
            self._cb_planes = {
                "X": QCheckBox("Sagittal (Flex/Ext)"),
                "Y": QCheckBox("Frontal (Ab/Adduction)"),
                "Z": QCheckBox("Transverse (Int/Ext Rotation)"),
            }
            for cb in self._cb_planes.values():
                cb.setChecked(True)
                plane_lay.addWidget(cb)
            left_col.addWidget(plane_grp)

            # Variable types
            type_grp = QGroupBox("Variable Type")
            type_lay = QVBoxLayout(type_grp)
            self._cb_types = {
                "Angles":  QCheckBox("Angles"),
                "Moments": QCheckBox("Moments"),
                "Powers":  QCheckBox("Powers"),
                "Forces":  QCheckBox("Forces"),
            }
            for key, cb in self._cb_types.items():
                cb.setChecked(key in ("Angles", "Moments", "Powers"))
                type_lay.addWidget(cb)
            left_col.addWidget(type_grp)
            left_col.addStretch(1)

            # Joints
            joint_grp = QGroupBox("Joint / Segment")
            joint_lay = QVBoxLayout(joint_grp)

            qs_row = QHBoxLayout()
            all_btn  = QPushButton("Select All")
            none_btn = QPushButton("Select None")
            std_btn  = QPushButton("Lower-limb only")
            qs_row.addWidget(all_btn)
            qs_row.addWidget(none_btn)
            qs_row.addWidget(std_btn)
            qs_row.addStretch()
            joint_lay.addLayout(qs_row)

            self._cb_joints: dict[str, QCheckBox] = {}
            grid = QGridLayout()
            grid.setHorizontalSpacing(20)
            joint_lay.addLayout(grid)
            cols = 2
            for i, name in enumerate(_DEFAULT_JOINTS):
                cb = QCheckBox(name)
                cb.setChecked(name in ("Hip", "Knee", "Ankle", "Pelvis"))
                self._cb_joints[name] = cb
                grid.addWidget(cb, i // cols, i % cols)

            custom_lbl = QLabel("Custom")
            custom_lbl.setStyleSheet(
                f"color: {PALETTE['text_muted']}; font-size: 11px;")
            joint_lay.addSpacing(6)
            joint_lay.addWidget(custom_lbl)
            self._custom_edit = QLineEdit()
            self._custom_edit.setPlaceholderText("e.g. Calc, MTP, ForeFoot")
            joint_lay.addWidget(self._custom_edit)

            top.addWidget(joint_grp, stretch=2)

            def _set_all(checked: bool):
                for cb in self._cb_joints.values():
                    cb.setChecked(checked)

            def _lower_only():
                for name, cb in self._cb_joints.items():
                    cb.setChecked(name in ("Hip", "Knee", "Ankle",
                                            "Pelvis", "FootProgress"))

            all_btn.clicked.connect(lambda: _set_all(True))
            none_btn.clicked.connect(lambda: _set_all(False))
            std_btn.clicked.connect(_lower_only)

            # Multi-trial average (only meaningful with >1 input file)
            avg_grp = QGroupBox("Multi-trial average")
            avg_lay = QVBoxLayout(avg_grp)
            self._cb_make_avg = QCheckBox(
                "Also produce one combined average file across all input "
                "trials (pools all strides)")
            self._cb_make_avg.setChecked(self._multi_trial)
            self._cb_make_avg.setEnabled(self._multi_trial)
            avg_lay.addWidget(self._cb_make_avg)
            if not self._multi_trial:
                hint = QLabel(
                    "  ↪  Enabled automatically when 2+ stride files are "
                    "selected.")
                hint.setStyleSheet(
                    f"color: {PALETTE['text_muted']}; font-size: 11px;")
                avg_lay.addWidget(hint)
            layout.addWidget(avg_grp)

            # Healthy reference (optional, full-width)
            hr_grp = QGroupBox("Healthy Reference (optional)")
            hr_lay = QVBoxLayout(hr_grp)
            self._cb_use_healthy = QCheckBox(
                "Compare patient features to healthy reference")
            hr_lay.addWidget(self._cb_use_healthy)

            file_row = QHBoxLayout()
            file_row.addWidget(QLabel("Healthy file:"))
            self._healthy_edit = QLineEdit()
            self._healthy_edit.setPlaceholderText(
                "Path to a *_features.txt from a healthy participant")
            file_row.addWidget(self._healthy_edit, stretch=1)
            self._healthy_browse = QPushButton("Browse…")
            self._healthy_browse.clicked.connect(self._browse_healthy)
            file_row.addWidget(self._healthy_browse)
            hr_lay.addLayout(file_row)

            def _toggle_hr(state):
                on = bool(state)
                self._healthy_edit.setEnabled(on)
                self._healthy_browse.setEnabled(on)
            self._cb_use_healthy.stateChanged.connect(_toggle_hr)
            _toggle_hr(False)
            layout.addWidget(hr_grp)

            btns = QDialogButtonBox(
                QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
            btns.accepted.connect(self._on_accept)
            btns.rejected.connect(self.reject)
            layout.addWidget(btns)

        def _browse_healthy(self):
            f, _ = QFileDialog.getOpenFileName(
                self, "Select Healthy Reference (.txt)",
                self._healthy_edit.text().strip(),
                "Feature summary text (*.txt);;All files (*.*)")
            if f:
                self._healthy_edit.setText(f)

        def _load_settings(self):
            s = QSettings(self._SETTINGS_ORG, self._SETTINGS_APP)
            path = str(s.value("healthy_path", "") or "")
            use  = s.value("use_healthy", False)
            if isinstance(use, str):
                use = use.lower() in ("1", "true", "yes")
            self._healthy_edit.setText(path)
            self._cb_use_healthy.setChecked(bool(use))
            self._healthy_edit.setEnabled(bool(use))
            self._healthy_browse.setEnabled(bool(use))

        def _save_settings(self):
            s = QSettings(self._SETTINGS_ORG, self._SETTINGS_APP)
            s.setValue("healthy_path", self._healthy_edit.text().strip())
            s.setValue("use_healthy",  self._cb_use_healthy.isChecked())

        def _on_accept(self):
            if not any(cb.isChecked() for cb in self._cb_sides.values()):
                QMessageBox.warning(self, "Select Sides",
                                    "Pick at least one side."); return
            if not any(cb.isChecked() for cb in self._cb_planes.values()):
                QMessageBox.warning(self, "Select Planes",
                                    "Pick at least one plane."); return
            if not any(cb.isChecked() for cb in self._cb_types.values()):
                QMessageBox.warning(self, "Select Types",
                                    "Pick at least one variable type."); return
            joints_picked = any(cb.isChecked() for cb in self._cb_joints.values())
            custom_text = self._custom_edit.text().strip()
            if not joints_picked and not custom_text:
                QMessageBox.warning(self, "Select Joints",
                                    "Pick at least one joint or custom pattern.")
                return
            if self._cb_use_healthy.isChecked():
                p = self._healthy_edit.text().strip()
                if not p:
                    QMessageBox.warning(self, "Healthy Reference",
                                        "Pick a healthy .txt file or uncheck "
                                        "the comparison option."); return
                if not Path(p).exists():
                    QMessageBox.warning(self, "Healthy Reference",
                                        f"File not found:\n{p}"); return
            self._save_settings()
            self.accept()

        def filter_spec(self) -> FilterSpec:
            sides   = {k for k, cb in self._cb_sides.items() if cb.isChecked()}
            planes  = {k for k, cb in self._cb_planes.items() if cb.isChecked()}
            types_  = {k for k, cb in self._cb_types.items() if cb.isChecked()}
            joints  = {k for k, cb in self._cb_joints.items() if cb.isChecked()}
            custom  = [s for s in self._custom_edit.text().split(",")
                        if s.strip()]
            healthy = (self._healthy_edit.text().strip()
                       if self._cb_use_healthy.isChecked() else None)
            return FilterSpec(sides=sides or None,
                              joints=joints or None,
                              plane_codes=planes or None,
                              data_types=types_ or None,
                              custom_joints=custom,
                              healthy_path=healthy)

        def make_average(self) -> bool:
            return self._cb_make_avg.isChecked() and self._cb_make_avg.isEnabled()

    # Worker
    class FeatureWorker(QThread):
        progress    = Signal(str, int)
        file_status = Signal(str, str, str)
        log_msg     = Signal(str)
        finished    = Signal(int, list)

        def __init__(self, files: list, out_dir,
                     filter_spec: FilterSpec | None = None,
                     make_average: bool = False):
            super().__init__()
            self.files        = files
            self.out_dir      = out_dir
            self.filter_spec  = filter_spec
            self.make_average = make_average

        def run(self):
            total   = len(self.files)
            success = 0
            errors  = []
            successful_paths: list[Path] = []

            if self.filter_spec is not None:
                self.log_msg.emit(
                    f"Filter: {self.filter_spec.describe()}")
            if self.make_average and total > 1:
                self.log_msg.emit(
                    f"Multi-trial average: ENABLED  "
                    f"({total} input file(s) → 1 combined average)")

            # ── Per-file pass ─────────────────────────────────────────────
            for i, xlsx_path in enumerate(self.files):
                # Reserve last 15% for the averaged pass if needed
                cap = 85 if (self.make_average and total > 1) else 100
                self.progress.emit(
                    f"[{i+1}/{total}]  {xlsx_path.name}",
                    int(cap * i / total))
                self.file_status.emit(str(xlsx_path), "Running…", "running")

                out_path = self.out_dir / f"{xlsx_path.stem}_features.xlsx"

                def _cb(msg, p=xlsx_path):
                    self.log_msg.emit(f"    {msg}")

                try:
                    self.log_msg.emit(
                        f"\n{'─'*50}\n{xlsx_path.name}\n{'─'*50}")
                    result = extract_gait_features(
                        xlsx_path, out_path,
                        filter_spec=self.filter_spec,
                        status_cb=_cb)
                    success += 1
                    successful_paths.append(xlsx_path)
                    n_rows = result["n_rows"]
                    nR = result["n_strides_R"]
                    nL = result["n_strides_L"]
                    self.file_status.emit(
                        str(xlsx_path),
                        f"✓ {n_rows} rows  (R:{nR}, L:{nL})", "ok")
                    self.log_msg.emit(
                        f"  ✓  {xlsx_path.name}  →  {out_path.name}")
                    txt = result.get("txt_path", "")
                    if txt:
                        self.log_msg.emit(
                            f"  ✓  LLM text  →  {Path(txt).name}")
                except Exception as exc:
                    import traceback
                    tb = traceback.format_exc()
                    errors.append((xlsx_path.name, str(exc)))
                    self.file_status.emit(
                        str(xlsx_path), f"✗ {exc}", "err")
                    self.log_msg.emit(
                        f"  ✗  {xlsx_path.name}:  {exc}\n{tb}")

            # ── Multi-trial average pass ──────────────────────────────────
            if self.make_average and len(successful_paths) >= 2:
                self.progress.emit("Building multi-trial average…", 88)
                self.log_msg.emit(
                    f"\n{'═'*50}\n"
                    f"MULTI-TRIAL AVERAGE  ({len(successful_paths)} trials)\n"
                    f"{'═'*50}")
                stem = _suggest_average_stem(successful_paths)
                avg_path = self.out_dir / f"{stem}.xlsx"

                def _cb(msg):
                    self.log_msg.emit(f"    {msg}")

                try:
                    res = extract_features_averaged(
                        successful_paths, avg_path,
                        filter_spec=self.filter_spec,
                        status_cb=_cb)
                    self.log_msg.emit(
                        f"  ✓  Average Excel  →  {avg_path.name}")
                    txt = res.get("txt_path", "")
                    if txt:
                        self.log_msg.emit(
                            f"  ✓  Average LLM text  →  {Path(txt).name}")
                except Exception as exc:
                    import traceback
                    tb = traceback.format_exc()
                    errors.append(("[multi-trial average]", str(exc)))
                    self.log_msg.emit(
                        f"  ✗  Multi-trial average failed:  {exc}\n{tb}")
            elif self.make_average and len(successful_paths) < 2:
                self.log_msg.emit(
                    "  ⚠  Multi-trial average skipped — "
                    "need at least 2 successful files.")

            self.finished.emit(success, errors)

    class FeatureExtractorTab(QWidget):

        def __init__(self, parent=None):
            super().__init__(parent)
            self._files: list[Path] = []
            self._out_dir: Path | None = None
            self._worker = None
            self._build()

        def _build(self):
            layout = QVBoxLayout(self)
            layout.setContentsMargins(0, 0, 0, 0)
            layout.setSpacing(0)

            from GaitSharing_ui import (make_page_header, make_accent_btn,
                                        make_console_log)

            layout.addWidget(make_page_header(
                "Gait Feature Extractor",
                "Extract ML features from normalised stride data "
                "(max, min, mean, range, max@, min@)"))

            content = QWidget()
            clay = QVBoxLayout(content)
            clay.setContentsMargins(16, 12, 16, 12)
            clay.setSpacing(10)

            file_group = QGroupBox(
                "Input Stride Files (from Stride Analysis)")
            fglay = QVBoxLayout(file_group)

            btn_row = QHBoxLayout()
            add_btn = make_accent_btn("+ Add Stride Excel Files")
            add_btn.clicked.connect(self._add_files)
            btn_row.addWidget(add_btn)
            rm_btn = QPushButton("Remove Selected")
            rm_btn.clicked.connect(self._remove_sel)
            btn_row.addWidget(rm_btn)
            clr_btn = QPushButton("Clear All")
            clr_btn.clicked.connect(self._clear)
            btn_row.addWidget(clr_btn)
            btn_row.addStretch()
            fglay.addLayout(btn_row)

            self.file_table = QTableWidget()
            self.file_table.setColumnCount(3)
            self.file_table.setHorizontalHeaderLabels(
                ["Filename", "Folder", "Status"])
            self.file_table.setAlternatingRowColors(True)
            self.file_table.setSelectionBehavior(
                QAbstractItemView.SelectRows)
            self.file_table.setSelectionMode(
                QAbstractItemView.ExtendedSelection)
            self.file_table.verticalHeader().setVisible(False)
            self.file_table.setEditTriggers(
                QAbstractItemView.NoEditTriggers)
            self.file_table.horizontalHeader().setStretchLastSection(True)
            self.file_table.setColumnWidth(0, 240)
            self.file_table.setColumnWidth(1, 340)
            self.file_table.setMaximumHeight(140)
            fglay.addWidget(self.file_table)

            self._file_count_lbl = QLabel("No files selected.")
            self._file_count_lbl.setStyleSheet(
                f"color: {PALETTE['text_muted']}; font-size: 12px;")
            fglay.addWidget(self._file_count_lbl)
            clay.addWidget(file_group)

            out_group = QGroupBox(
                "Output Folder (one _features.xlsx + .txt per input file"
                "; plus one combined average when 2+ files are loaded)")
            oglay = QHBoxLayout(out_group)
            self._out_lbl = QLabel("No folder selected.")
            self._out_lbl.setStyleSheet(
                f"color: {PALETTE['text_muted']};")
            oglay.addWidget(self._out_lbl, stretch=1)
            brw_btn = QPushButton("Browse…")
            brw_btn.clicked.connect(self._browse_out)
            oglay.addWidget(brw_btn)
            clay.addWidget(out_group)

            info = QLabel(
                "ℹ  Reads normalised sheets from stride analysis output.\n"
                "     A pop-up will let you select sides / joints / planes / "
                "variable types.\n"
                "     Per-file output: one Clinical_Summary sheet + companion "
                ".txt for AI.\n"
                "     With 2+ input files: also produces a single "
                "*_AVERAGE_features.xlsx + .txt that pools all strides "
                "across trials.")
            info.setWordWrap(True)
            info.setStyleSheet(
                f"color: {PALETTE['text_muted']}; font-size: 12px; "
                f"padding: 4px 0;")
            clay.addWidget(info)

            self._run_btn = make_accent_btn("▶   Extract Features")
            self._run_btn.clicked.connect(self._run)
            clay.addWidget(self._run_btn, alignment=Qt.AlignLeft)

            prog_group = QGroupBox("Progress")
            pglay = QVBoxLayout(prog_group)
            self._status_lbl = QLabel("Ready.")
            pglay.addWidget(self._status_lbl)
            self._pbar = QProgressBar()
            self._pbar.setMaximum(100)
            pglay.addWidget(self._pbar)
            clay.addWidget(prog_group)

            log_group = QGroupBox("Log")
            lglay = QVBoxLayout(log_group)
            self._log_widget = make_console_log()
            self._log_widget.setMinimumHeight(100)
            lglay.addWidget(self._log_widget)
            clay.addWidget(log_group, stretch=1)

            scroll = QScrollArea()
            scroll.setWidgetResizable(True)
            scroll.setWidget(content)
            scroll.setStyleSheet("QScrollArea { border: none; }")
            layout.addWidget(scroll)

        def _log(self, msg: str):
            self._log_widget.append(msg)

        def _refresh_file_list(self):
            self.file_table.setRowCount(len(self._files))
            for i, p in enumerate(self._files):
                self.file_table.setItem(
                    i, 0, QTableWidgetItem(p.name))
                self.file_table.setItem(
                    i, 1, QTableWidgetItem(str(p.parent)))
                item = QTableWidgetItem("Pending")
                item.setForeground(QColor(PALETTE["text_muted"]))
                self.file_table.setItem(i, 2, item)
            n = len(self._files)
            self._file_count_lbl.setText(
                f"{n} file{'s' if n != 1 else ''} selected."
                + (f"   →  +1 combined average will be produced."
                   if n >= 2 else "")
                if n else "No files selected.")

        def _set_file_status(self, path_str, text, status_type):
            for row in range(self.file_table.rowCount()):
                item0 = self.file_table.item(row, 0)
                item1 = self.file_table.item(row, 1)
                if item0 and item1:
                    full = str(Path(item1.text()) / item0.text())
                    if full == path_str:
                        status_item = QTableWidgetItem(text)
                        color_map = {
                            "running": PALETTE["warning"],
                            "ok":      PALETTE["success"],
                            "err":     "#C42B1C",
                        }
                        status_item.setForeground(
                            QColor(color_map.get(
                                status_type, PALETTE["text"])))
                        self.file_table.setItem(row, 2, status_item)
                        break

        def _add_files(self):
            files, _ = QFileDialog.getOpenFileNames(
                self, "Select Stride Excel Files",
                filter="Excel files (*.xlsx);;All files (*.*)")
            for raw in files:
                p = Path(raw)
                if p not in self._files:
                    self._files.append(p)
            self._refresh_file_list()

        def _remove_sel(self):
            rows = sorted(
                set(idx.row()
                    for idx in self.file_table.selectedIndexes()),
                reverse=True)
            for row in rows:
                if row < len(self._files):
                    self._files.pop(row)
            self._refresh_file_list()

        def _clear(self):
            self._files.clear()
            self._refresh_file_list()

        def _browse_out(self):
            d = QFileDialog.getExistingDirectory(
                self, "Select Output Folder")
            if d:
                self._out_dir = Path(d)
                self._out_lbl.setText(str(self._out_dir))

        def _run(self):
            if not self._files:
                QMessageBox.warning(
                    self, "No Files",
                    "Add at least one _strides.xlsx file.")
                return
            if not self._out_dir:
                QMessageBox.critical(
                    self, "No Output Folder",
                    "Select an output folder.")
                return

            multi = len(self._files) >= 2
            dlg = FilterSelectionDialog(self, multi_trial=multi)
            if dlg.exec() != QDialog.Accepted:
                self._log("  (filter dialog cancelled — run aborted)")
                return
            spec = dlg.filter_spec()
            make_avg = dlg.make_average()

            self._run_btn.setEnabled(False)
            self._pbar.setValue(0)

            self._worker = FeatureWorker(
                self._files[:], self._out_dir,
                filter_spec=spec,
                make_average=make_avg)
            self._worker.progress.connect(
                lambda msg, pct: (
                    self._status_lbl.setText(msg),
                    self._pbar.setValue(pct)))
            self._worker.file_status.connect(self._set_file_status)
            self._worker.log_msg.connect(self._log)
            self._worker.finished.connect(self._on_finished)
            self._worker.start()

        def _on_finished(self, success, errors):
            total = len(self._files)
            self._pbar.setValue(100)
            self._status_lbl.setText(
                f"Done — {success}/{total} files, "
                f"{len(errors)} error(s).")
            self._run_btn.setEnabled(True)

            if errors:
                QMessageBox.warning(
                    self, "Done",
                    f"{success} OK, {len(errors)} failed.\n\n" +
                    "\n".join(f"• {n}: {e}"
                              for n, e in errors[:5]))
            else:
                QMessageBox.information(
                    self, "Done",
                    f"✓  {success} file(s) saved to:\n"
                    f"{self._out_dir}")

#  CLI

if __name__ == "__main__":
    import sys
    args = sys.argv[1:]
    if not args:
        print("Usage:")
        print("  Single trial:")
        print("    python GaitSharing_features.py <strides.xlsx> [output.xlsx]")
        print("  Multi-trial average (2+ inputs, last arg = output):")
        print("    python GaitSharing_features.py <s1.xlsx> <s2.xlsx> ... "
              "<output_average.xlsx>")
        sys.exit(1)

    if len(args) == 1:
        inp = Path(args[0])
        out = inp.with_name(f"{inp.stem}_features{inp.suffix}")
        result = extract_gait_features(inp, out, status_cb=print)
        print(f"\nResult: {result}")
    elif len(args) == 2:
        inp = Path(args[0])
        out = Path(args[1])
        result = extract_gait_features(inp, out, status_cb=print)
        print(f"\nResult: {result}")
    else:
        # 3+ args: treat last as output, the rest as inputs
        inputs = [Path(a) for a in args[:-1]]
        out    = Path(args[-1])
        print(f"Multi-trial mode: {len(inputs)} input(s) → {out.name}")
        result = extract_features_averaged(inputs, out, status_cb=print)
        print(f"\nResult: {result}")
